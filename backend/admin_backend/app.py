import os, uuid, json, datetime, io, csv, math, time, hmac, hashlib, base64, logging, secrets, re, threading
from urllib.parse import quote_plus
from flask import Flask, request, jsonify, send_from_directory, send_file, Blueprint, current_app
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.utils import secure_filename
import jwt
import bcrypt
import pymysql
import qrcode
import openpyxl
from admin_backend.extensions import db
import requests
import email_service

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

MYSQL_HOST = os.environ.get('MYSQL_HOST', '127.0.0.1')
MYSQL_PORT = int(os.environ.get('MYSQL_PORT', '3306'))
MYSQL_USER = os.environ.get('MYSQL_USER', 'meal_app')
MYSQL_PASSWORD = os.environ.get('MYSQL_PASSWORD', 'Admin@RKMVC2')
MYSQL_DATABASE = os.environ.get('MYSQL_DATABASE', 'rkmvc_mealflow_db')

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# app = Flask(__name__, static_folder='dist', static_url_path='')
admin_bp = Blueprint('admin_bp', __name__)
# app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB max upload

# cors_origins = os.environ.get('CORS_ORIGINS', '*')
# if cors_origins == '*':
#     CORS(app)
# else:
#     CORS(app, origins=cors_origins.split(','))



limiter = Limiter(key_func=get_remote_address, default_limits=["5000 per day", "1000 per hour"])

JWT_SECRET = os.environ.get('JWT_SECRET', 'change-this-to-a-random-64-char-string')
QR_HMAC_SECRET = os.environ.get('QR_HMAC_SECRET', 'change-this-to-another-random-64-char-string')
TOKEN_EXPIRY_DEFAULT = int(os.environ.get('TOKEN_EXPIRY_DEFAULT', '60'))
CREDENTIALS_CSV_PATH = os.environ.get('CREDENTIALS_CSV_PATH', 'student_credentials.csv')
REGISTRATION_BACKEND_URL = os.environ.get('REGISTRATION_BACKEND_URL', 'http://localhost:5000')

# Quiet the favicon.ico noise
import warnings
warnings.filterwarnings('ignore', message='.*favicon.ico.*')

from datetime import timedelta, date
def serialize_row(row):
    if isinstance(row, dict):
        return {k: serialize_row(v) for k, v in row.items()}
    if isinstance(row, (list, tuple)):
        return [serialize_row(v) for v in row]
    if isinstance(row, timedelta):
        total = int(row.total_seconds())
        h, rem = divmod(total, 3600)
        m, s = divmod(rem, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"
    if isinstance(row, (date, datetime.time)):
        return row.isoformat()
    return row

# app.config['SQLALCHEMY_DATABASE_URI'] = (
#     f"mysql+pymysql://{MYSQL_USER}:{MYSQL_PASSWORD}@{MYSQL_HOST}:{MYSQL_PORT}/{MYSQL_DATABASE}"
# )
# app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# db.init_app(app)

def sani(msg):
    """Return a safe generic message, log the real one."""
    if msg:
        logger.error(msg)
    return "An internal error occurred. Please try again."

def get_db():
    return _get_mysql_connection()

def _get_mysql_connection():
    target_host = os.environ.get('MYSQL_HOST', os.environ.get('DB_HOST', '127.0.0.1'))
    hosts_to_try = [target_host, 'db', '127.0.0.1']
    
    for host_candidate in dict.fromkeys(hosts_to_try):
        try:
            return pymysql.connect(host=host_candidate, port=MYSQL_PORT, user=MYSQL_USER,
                                   password=MYSQL_PASSWORD, database=MYSQL_DATABASE,
                                   charset='utf8mb4', autocommit=True, cursorclass=pymysql.cursors.DictCursor)
        except Exception:
            continue

    root_passwords = [
        os.environ.get('MYSQL_ROOT_PASSWORD', 'root_password_secure'),
        'AkashPillai@123',
        'root_password_secure',
        ''
    ]
    for host_candidate in dict.fromkeys(hosts_to_try):
        for r_pass in root_passwords:
            try:
                bootstrap = pymysql.connect(host=host_candidate, port=MYSQL_PORT, user='root',
                                            password=r_pass, charset='utf8mb4', autocommit=True)
                with bootstrap.cursor() as cur:
                    cur.execute(f"CREATE DATABASE IF NOT EXISTS `{MYSQL_DATABASE}` CHARACTER SET utf8mb4")
                    try:
                        cur.execute(f"CREATE USER IF NOT EXISTS '{MYSQL_USER}'@'%' IDENTIFIED BY '{MYSQL_PASSWORD}'")
                        cur.execute(f"GRANT ALL PRIVILEGES ON `{MYSQL_DATABASE}`.* TO '{MYSQL_USER}'@'%'")
                        cur.execute(f"CREATE USER IF NOT EXISTS '{MYSQL_USER}'@'localhost' IDENTIFIED BY '{MYSQL_PASSWORD}'")
                        cur.execute(f"GRANT ALL PRIVILEGES ON `{MYSQL_DATABASE}`.* TO '{MYSQL_USER}'@'localhost'")
                        cur.execute("FLUSH PRIVILEGES")
                    except Exception:
                        pass
                bootstrap.close()

                try:
                    return pymysql.connect(host=host_candidate, port=MYSQL_PORT, user=MYSQL_USER,
                                           password=MYSQL_PASSWORD, database=MYSQL_DATABASE,
                                           charset='utf8mb4', autocommit=True, cursorclass=pymysql.cursors.DictCursor)
                except Exception:
                    return pymysql.connect(host=host_candidate, port=MYSQL_PORT, user='root',
                                           password=r_pass, database=MYSQL_DATABASE,
                                           charset='utf8mb4', autocommit=True, cursorclass=pymysql.cursors.DictCursor)
            except Exception:
                continue

    raise pymysql.err.OperationalError(2003, f"Can't connect to MySQL server on {target_host}")

def _ensure_student_meals_columns(cur):
    """Ensures required columns exist in student_meals table for media ingestion & student migration."""
    student_meals_cols = [
        ('username', 'VARCHAR(50) NULL'),
        ('email', 'VARCHAR(100) NULL'),
        ('password_hash', 'VARCHAR(255) NULL'),
        ('degree_year', 'VARCHAR(50) NULL'),
        ('mobile_no', 'VARCHAR(50) NULL'),
        ('qr_secret', 'VARCHAR(64) NULL'),
        ('image_url', 'VARCHAR(512) NULL'),
        ('image_path', 'VARCHAR(512) NULL'),
        ('student_image_path', 'VARCHAR(512) NULL'),
        ('last_served_date', 'DATE NULL'),
        ('created_at', 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
    ]
    for col, col_def in student_meals_cols:
        try:
            cur.execute(f"ALTER TABLE student_meals ADD COLUMN {col} {col_def}")
        except Exception:
            pass

    # Drop unneeded academic_year, income_proof_path and signature_path columns from student_meals if present
    for old_col in ('academic_year', 'income_proof_path', 'signature_path'):
        try:
            cur.execute(f"ALTER TABLE student_meals DROP COLUMN `{old_col}`")
        except Exception:
            pass

    try:
        cur.execute("ALTER TABLE meal_tokens ADD COLUMN cached_student_name VARCHAR(100) NULL")
    except Exception:
        pass

def _ensure_meal_registrations_columns(cur):
    """Ensures required columns exist in meal_registrations table."""
    cols = [
        ('date_of_birth', 'VARCHAR(50) NULL AFTER dob_age'),
        ('age', 'INT NULL AFTER date_of_birth')
    ]
    for col, col_def in cols:
        try:
            cur.execute(f"ALTER TABLE meal_registrations ADD COLUMN {col} {col_def}")
        except Exception:
            pass

def _ensure_indexes(cur):
    """Safely bootstraps SQL indexes matching physical table schema (GAP 2)."""
    index_targets = [
        ('student_meals', 'idx_student_meals_reg', ['student_id', 'register_number', 'registration_number']),
        ('meal_tokens', 'idx_meal_tokens_reg', ['student_id', 'registration_number', 'register_number']),
        ('meal_distribution_log', 'idx_log_reg_num', ['student_id', 'register_number', 'registration_number', 'token_id'])
    ]
    for table_name, index_name, possible_cols in index_targets:
        try:
            cur.execute(f"DESCRIBE `{table_name}`")
            existing_cols = [c['Field'] for c in cur.fetchall()]
            matched_col = next((col for col in possible_cols if col in existing_cols), None)
            if matched_col:
                cur.execute(f"SHOW INDEX FROM `{table_name}` WHERE Key_name = %s", (index_name,))
                if not cur.fetchone():
                    cur.execute(f"CREATE INDEX `{index_name}` ON `{table_name}`(`{matched_col}`)")
        except Exception as idx_err:
            logger.warning("Notice creating index %s on %s: %s", index_name, table_name, idx_err)

def _ensure_tables(conn):
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id VARCHAR(50) PRIMARY KEY, username VARCHAR(50) UNIQUE NOT NULL,
                email VARCHAR(100) UNIQUE NOT NULL, password_hash VARCHAR(255) NOT NULL,
                role ENUM('admin','approval_staff','canteen_staff') NOT NULL DEFAULT 'approval_staff',
                display_name VARCHAR(100) NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_meals (
                student_id VARCHAR(50) PRIMARY KEY,
                username VARCHAR(50) UNIQUE NULL,
                email VARCHAR(100) NULL,
                password_hash VARCHAR(255) NULL,
                name VARCHAR(100) NOT NULL,
                grade_section VARCHAR(100) NOT NULL,
                forenoon_meal TINYINT(1) DEFAULT 1, afternoon_meal TINYINT(1) DEFAULT 1,
                last_served_date DATE NULL, qr_secret VARCHAR(64) NULL,
                image_url VARCHAR(512) NULL, image_path VARCHAR(512) NULL,
                student_image_path VARCHAR(512) NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        # Migrations for existing databases
        try:
            cur.execute("DELETE FROM users WHERE role NOT IN ('admin','approval_staff','canteen_staff')")
            cur.execute("ALTER TABLE users MODIFY COLUMN role ENUM('admin','approval_staff','canteen_staff') NOT NULL DEFAULT 'approval_staff'")
        except Exception:
            pass

        try:
            cur.execute("ALTER TABLE users DROP COLUMN student_id")
        except Exception:
            pass

        _ensure_student_meals_columns(cur)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS meal_registrations (
                registration_id VARCHAR(50) PRIMARY KEY,
                app_no VARCHAR(50) NULL,
                student_name VARCHAR(100) NOT NULL,
                dob_age VARCHAR(50) NULL,
                date_of_birth VARCHAR(50) NULL,
                age INT NULL,
                course VARCHAR(100) NULL,
                department VARCHAR(100) NULL,
                degree_year VARCHAR(20) NULL,
                dept_number VARCHAR(50) NULL,
                mobile_no VARCHAR(20) NULL,
                email VARCHAR(100) NULL,
                father_name VARCHAR(100) NULL,
                father_occupation VARCHAR(100) NULL,
                forenoon_meal TINYINT(1) DEFAULT 1,
                afternoon_meal TINYINT(1) DEFAULT 1,
                annual_income VARCHAR(50) NULL,
                distance_km VARCHAR(50) NULL,
                permanent_address TEXT NULL,
                permanent_pin VARCHAR(20) NULL,
                local_address TEXT NULL,
                local_pin VARCHAR(20) NULL,
                landline VARCHAR(50) NULL,
                employment_type VARCHAR(50) NULL,
                religion VARCHAR(50) NULL,
                community VARCHAR(50) NULL,
                last_year_id VARCHAR(50) NULL,
                student_photo_url VARCHAR(512) NULL,
                applicant_signature_url VARCHAR(512) NULL,
                income_proof_url VARCHAR(512) NULL,
                generated_pdf_url VARCHAR(512) NULL,
                status ENUM('pending','approved','rejected') DEFAULT 'pending',
                submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_reg_status (status),
                INDEX idx_reg_dept (dept_number),
                INDEX idx_reg_mobile (mobile_no)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        _ensure_meal_registrations_columns(cur)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS meal_windows (
                id INT AUTO_INCREMENT PRIMARY KEY,
                meal_type ENUM('forenoon','afternoon') NOT NULL,
                day_of_week TINYINT NULL, start_time TIME NOT NULL, end_time TIME NOT NULL,
                expiry_minutes INT DEFAULT 30, is_active TINYINT(1) DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        try:
            cur.execute("SHOW COLUMNS FROM meal_windows LIKE 'grace_minutes'")
            if cur.fetchone():
                cur.execute("ALTER TABLE meal_windows CHANGE COLUMN grace_minutes expiry_minutes INT DEFAULT 30")
        except Exception:
            pass
        try:
            cur.execute("UPDATE meal_windows SET expiry_minutes = 30 WHERE expiry_minutes = 15")
        except Exception:
            pass
        cur.execute("""
            CREATE TABLE IF NOT EXISTS meal_tokens (
                id INT AUTO_INCREMENT PRIMARY KEY,
                token_uid VARCHAR(50) UNIQUE NOT NULL,
                student_id VARCHAR(50) NOT NULL,
                meal_type ENUM('forenoon','afternoon') NOT NULL,
                status ENUM('awaiting_scan','staff_verified','approved','rejected','token_issued','redeemed','expired') DEFAULT 'awaiting_scan',
                scanned_by VARCHAR(50) NULL, scanned_at TIMESTAMP NULL,
                approved_by VARCHAR(50) NULL, approved_at TIMESTAMP NULL,
                reject_reason VARCHAR(255) NULL,
                token_qr_data TEXT NULL, token_issued_at TIMESTAMP NULL,
                redeemed_by VARCHAR(50) NULL, redeemed_at TIMESTAMP NULL,
                expiry_time TIMESTAMP NULL, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_token_uid (token_uid), INDEX idx_student (student_id),
                INDEX idx_status (status),
                FOREIGN KEY (student_id) REFERENCES student_meals(student_id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        try:
            cur.execute("""
                ALTER TABLE meal_tokens MODIFY COLUMN status 
                ENUM('active','awaiting_scan','staff_verified','approved','rejected','token_issued','redeemed','claimed','expired') 
                DEFAULT 'active'
            """)
        except Exception:
            pass
        cur.execute("""
            CREATE TABLE IF NOT EXISTS scan_audit_log (
                id BIGINT AUTO_INCREMENT PRIMARY KEY,
                scanner_id VARCHAR(50) NOT NULL,
                scanner_role ENUM('approval_staff','canteen_staff') NOT NULL,
                scan_type ENUM('student_id_qr','token_qr') NOT NULL,
                payload TEXT NOT NULL, student_id VARCHAR(50) NULL, token_uid VARCHAR(50) NULL,
                result ENUM('success','invalid_token','already_redeemed','expired','out_of_window',
                    'duplicate_meal','not_eligible','generation_disabled','invalid_signature','not_found') NOT NULL,
                detail TEXT NULL, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_scanner (scanner_id), INDEX idx_created (created_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS import_logs (
                id VARCHAR(50) PRIMARY KEY, filename VARCHAR(255) NOT NULL,
                records_imported INT DEFAULT 0,
                status ENUM('SUCCESS','FAILED','PARTIAL') DEFAULT 'SUCCESS',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS export_logs (
                id VARCHAR(50) PRIMARY KEY, filename VARCHAR(255) NOT NULL,
                records_exported INT DEFAULT 0, format ENUM('csv','excel','json') DEFAULT 'csv',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id VARCHAR(50) PRIMARY KEY, username VARCHAR(50) NOT NULL,
                action VARCHAR(50) NOT NULL, table_name VARCHAR(50) NOT NULL,
                details TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS meal_distribution_log (
                log_id INT AUTO_INCREMENT PRIMARY KEY,
                token_id VARCHAR(50) NOT NULL,
                session_type VARCHAR(20) NOT NULL,
                status VARCHAR(20) NOT NULL DEFAULT 'Distributed',
                served_by VARCHAR(50) NULL,
                served_at DATETIME NOT NULL,
                timestamp DATETIME NULL,
                UNIQUE KEY uk_token_id (token_id),
                CONSTRAINT fk_dist_meal_tokens FOREIGN KEY (token_id) REFERENCES meal_tokens(token_uid) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS app_state (
                id INT PRIMARY KEY, data JSON NOT NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS system_roles (
                id INT AUTO_INCREMENT PRIMARY KEY,
                role_name VARCHAR(50) UNIQUE NOT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        cur.execute("SELECT COUNT(*) as c FROM system_roles")
        if cur.fetchone()['c'] == 0:
            for r in ['admin','approval_staff','canteen_staff','student']:
                cur.execute("INSERT IGNORE INTO system_roles (role_name) VALUES (%s)", (r,))
        cur.execute("SELECT COUNT(*) as c FROM users WHERE role = 'admin'")
        if cur.fetchone()['c'] == 0:
            _seed_data(conn)
        # Always guarantee the default admin account exists with the known password
        import bcrypt as _bcrypt
        _default_pw_hash = _bcrypt.hashpw(b'adminpassword', _bcrypt.gensalt()).decode()
        cur.execute("""
            INSERT INTO users (id, username, email, password_hash, role, display_name)
            VALUES ('usr_admin_default', 'rkmvc_admin', 'admin@rkmvc.edu', %s, 'admin', 'RKMVC Administrator')
            ON DUPLICATE KEY UPDATE password_hash = VALUES(password_hash)
        """, (_default_pw_hash,))
        _sync_approved_registrations_to_student_meals(conn)
        
        # Safe Index Bootstrapping (GAP 2)
        try:
            _ensure_indexes(cur)
        except Exception as idx_e:
            logger.warning("Notice running index bootstrap: %s", idx_e)
    # Runs on every startup (idempotent): makes sure the app_state JSON blob
    # exists. This blob backs the generic "virtual tables" API used by the
    # admin Database/Export/Data-Tools pages (e.g. meal_distribution_log).
    # In the Docker setup this row comes from mysql/init/002-seed.sql, which
    # only runs via MySQL's container-init mechanism — so for a plain
    # `python app.py` / non-Docker run we create it here instead.
    _ensure_app_state(conn)

def _format_degree_year(val):
    if not val or str(val).strip() == '':
        return ''
    s = str(val).strip()
    s_lower = s.lower()
    if s_lower == 'enrolled':
        return ''
    if '1' in s_lower or 'first' in s_lower or s_lower == 'i':
        return '1st Year'
    elif '2' in s_lower or 'second' in s_lower or s_lower == 'ii':
        return '2nd Year'
    elif '3' in s_lower or 'third' in s_lower or s_lower == 'iii':
        return '3rd Year'
    elif 'graduat' in s_lower or 'complet' in s_lower:
        return 'Graduated'
    elif 'year' in s_lower:
        return s.title()
    else:
        return f"{s} Year" if not s.endswith('Year') else s

def _sync_approved_registrations_to_student_meals(conn):
    """Makes sure all approved registrations in meal_registrations are inserted into student_meals."""
    try:
        from urllib.parse import quote_plus
        with conn.cursor() as cur:
            _ensure_student_meals_columns(cur)
            cur.execute("SELECT * FROM meal_registrations WHERE status = 'approved'")
            approved_regs = cur.fetchall()
            for r in approved_regs:
                sid = (r.get('dept_number') or r.get('last_year_id') or r.get('app_no') or r.get('registration_id') or '').strip()
                if not sid:
                    continue
                username = sid
                raw_pw = 'pass123'
                pw_hash = bcrypt.hashpw(raw_pw.encode('utf-8'), bcrypt.gensalt()).decode()
                display_name = r.get('student_name') or username
                student_email = r.get('email') or f"{username.lower()}@student.rkmvc"
                grade_sec = f"{r.get('course', '')} - {r.get('department', '')}".strip(' -') or 'B.Sc. Comp Sci'
                deg_year = _format_degree_year(r.get('degree_year') or r.get('year_of_degree'))
                mobile_num = r.get('mobile_no') or r.get('phone') or 'N/A'
                fn_meal = 1 if r.get('forenoon_meal') in [True, 1, 'true', '1', 'True'] else 0
                an_meal = 1 if r.get('afternoon_meal') in [True, 1, 'true', '1', 'True'] else 0
                qr_sec = _gen_qr_secret(sid)
                img_url = r.get('student_image_path') or r.get('student_photo_url') or f"https://ui-avatars.com/api/?name={quote_plus(display_name)}&background=random"
                img_path = r.get('student_image_path') or r.get('student_photo_url') or ''
                cur.execute("""
                    INSERT INTO student_meals (
                        student_id, username, email, password_hash, name, grade_section,
                        degree_year, mobile_no,
                        forenoon_meal, afternoon_meal, qr_secret, image_url, image_path,
                        student_image_path
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    ) ON DUPLICATE KEY UPDATE
                        username=VALUES(username),
                        email=VALUES(email),
                        name=VALUES(name),
                        grade_section=VALUES(grade_section),
                        degree_year=IF(student_meals.degree_year IS NULL OR student_meals.degree_year = '' OR student_meals.degree_year = 'Enrolled', VALUES(degree_year), student_meals.degree_year),
                        mobile_no=VALUES(mobile_no),
                        forenoon_meal=student_meals.forenoon_meal,
                        afternoon_meal=student_meals.afternoon_meal,
                        qr_secret=IF(student_meals.qr_secret IS NULL OR student_meals.qr_secret = '', VALUES(qr_secret), student_meals.qr_secret),
                        image_url=IF(student_meals.image_url IS NULL OR student_meals.image_url = '', VALUES(image_url), student_meals.image_url),
                        image_path=IF(student_meals.image_path IS NULL OR student_meals.image_path = '', VALUES(image_path), student_meals.image_path),
                        student_image_path=IF(student_meals.student_image_path IS NULL OR student_meals.student_image_path = '', VALUES(student_image_path), student_meals.student_image_path)
                """, (
                    sid, username, student_email, pw_hash, display_name, grade_sec,
                    deg_year, mobile_num,
                    fn_meal, an_meal, qr_sec, img_url, img_path,
                    img_path
                ))

            # Run one-time SQL migration to populate missing degree_year for existing student_meals rows from meal_registrations
            try:
                cur.execute("""
                    UPDATE student_meals sm
                    INNER JOIN meal_registrations mr 
                       ON LOWER(TRIM(sm.student_id)) = LOWER(TRIM(mr.dept_number))
                       OR LOWER(TRIM(sm.student_id)) = LOWER(TRIM(mr.app_no))
                    SET sm.degree_year = mr.degree_year
                    WHERE (sm.degree_year IS NULL OR sm.degree_year = '' OR sm.degree_year = 'Enrolled')
                      AND (mr.degree_year IS NOT NULL AND mr.degree_year != '');
                """)
                conn.commit()
            except Exception as backfill_err:
                logger.warning("Notice backfilling degree_year: %s", backfill_err)
    except Exception as e:
        print("APPROVAL CRASH LOG (sync):", str(e), flush=True)
        logger.warning("Notice syncing approved registrations to student_meals: %s", e)

def _seed_data(conn):
    import bcrypt
    with conn.cursor() as cur:
        cur.execute("SET FOREIGN_KEY_CHECKS = 0")
        admin_pw = bcrypt.hashpw(b'adminpassword', bcrypt.gensalt()).decode()
        staff_pw = bcrypt.hashpw(b'staffpassword', bcrypt.gensalt()).decode()
        
        # 1. Operators (Admin & Staff ONLY)
        users = [
            ('usr_admin', 'admin', 'admin@example.com', admin_pw, 'admin', 'System Administrator'),
            ('usr_staff', 'staff', 'staff@example.com', staff_pw, 'approval_staff', 'Staff Operator'),
            ('usr_staff101', 'STAFF101', 'staff101@example.com', staff_pw, 'approval_staff', 'Morning Warden'),
            ('usr_staff102', 'STAFF102', 'staff102@example.com', staff_pw, 'approval_staff', 'Noon Warden'),
            ('usr_canteen01', 'CANTEEN01', 'canteen01@example.com', staff_pw, 'canteen_staff', 'Counter Terminal A'),
            ('usr_canteen02', 'CANTEEN02', 'canteen02@example.com', staff_pw, 'canteen_staff', 'Counter Terminal B'),
        ]
        for u in users:
            cur.execute("""
                INSERT INTO users (id, username, email, password_hash, role, display_name) 
                VALUES (%s, %s, %s, %s, %s, %s) 
                ON DUPLICATE KEY UPDATE password_hash=VALUES(password_hash), role=VALUES(role)
            """, u)

        cur.execute("SET FOREIGN_KEY_CHECKS = 1")

def _ensure_app_state(conn):
    """Creates/updates the app_state JSON blob (id=1)."""
    with conn.cursor() as cur:
        cur.execute("SELECT data FROM app_state WHERE id = 1")
        existing_row = cur.fetchone()
        if existing_row:
            try:
                db_data = json.loads(existing_row['data'])
                modified = False
                tables = db_data.setdefault('tables', {})
                pk_map = {
                    "users": "id",
                    "student_meals": "student_id",
                    "meal_tokens": "token_uid",
                    "meal_registrations": "registration_id",
                    "meal_distribution_log": "log_id",
                    "scan_audit_log": "id",
                    "import_logs": "id",
                    "export_logs": "id"
                }
                for tbl_name, pk_col_name in pk_map.items():
                    if tbl_name in tables:
                        cols = tables[tbl_name].get('columns', [])
                        if not any(col.get('primaryKey') for col in cols):
                            for col in cols:
                                if col.get('name') == pk_col_name or col.get('name') == 'id':
                                    col['primaryKey'] = True
                                    modified = True
                                    break
                if modified:
                    cur.execute("UPDATE app_state SET data = %s WHERE id = 1", (json.dumps(db_data, default=str, ensure_ascii=False),))
            except Exception as e:
                logger.warning("Notice updating app_state schema: %s", e)
            return

        cur.execute("SELECT id,username,email,role,display_name FROM users")
        user_rows = cur.fetchall()

        cur.execute("SELECT student_id,username,name,grade_section,forenoon_meal,afternoon_meal,last_served_date,image_url FROM student_meals ORDER BY student_id ASC")
        student_rows = cur.fetchall()
        for s in student_rows:
            s['forenoon_meal'] = bool(s['forenoon_meal'])
            s['afternoon_meal'] = bool(s['afternoon_meal'])
            if s.get('last_served_date') is not None:
                s['last_served_date'] = str(s['last_served_date'])

        tables = {
            "users": {
                "columns": [
                    {"name": "id", "type": "TEXT", "primaryKey": True}, {"name": "role", "type": "TEXT"},
                    {"name": "email", "type": "TEXT"}, {"name": "username", "type": "TEXT"},
                    {"name": "display_name", "type": "TEXT"}
                ],
                "rows": user_rows
            },
            "student_meals": {
                "columns": [
                    {"name": "student_id", "type": "TEXT", "primaryKey": True}, {"name": "username", "type": "TEXT"},
                    {"name": "name", "type": "TEXT"}, {"name": "grade_section", "type": "TEXT"}, 
                    {"name": "forenoon_meal", "type": "BOOLEAN"}, {"name": "afternoon_meal", "type": "BOOLEAN"}, 
                    {"name": "last_served_date", "type": "DATE"}, {"name": "image_url", "type": "TEXT"}
                ],
                "rows": student_rows
            },
            "meal_tokens": {
                "columns": [
                    {"name": "token_uid", "type": "TEXT"}, {"name": "student_id", "type": "TEXT"},
                    {"name": "meal_type", "type": "TEXT"}, {"name": "status", "type": "TEXT"},
                    {"name": "scanned_by", "type": "TEXT", "nullable": True}, {"name": "approved_by", "type": "TEXT", "nullable": True},
                    {"name": "redeemed_by", "type": "TEXT", "nullable": True}, {"name": "issued_at", "type": "DATETIME"},
                    {"name": "scanned_at", "type": "DATETIME", "nullable": True}, {"name": "approved_at", "type": "DATETIME", "nullable": True},
                    {"name": "redeemed_at", "type": "DATETIME", "nullable": True}, {"name": "expires_at", "type": "DATETIME"}
                ],
                "rows": []
            },
            "scan_audit_log": {
                "columns": [
                    {"name": "id", "type": "TEXT"}, {"name": "scanner_id", "type": "TEXT"},
                    {"name": "scanner_role", "type": "TEXT"}, {"name": "scan_type", "type": "TEXT"},
                    {"name": "result", "type": "TEXT"}, {"name": "student_id", "type": "TEXT"},
                    {"name": "token_uid", "type": "TEXT"}, {"name": "scanned_at", "type": "DATETIME"}
                ],
                "rows": []
            },
            "meal_distribution_log": {
                "columns": [
                    {"name": "log_id", "type": "TEXT"}, {"name": "student_id", "type": "TEXT"},
                    {"name": "session_type", "type": "TEXT"}, {"name": "status", "type": "TEXT"},
                    {"name": "served_by", "type": "TEXT"}, {"name": "served_at", "type": "DATETIME"}
                ],
                "rows": []
            },
            "meal_registrations": {
                "columns": [
                    {"name": "registration_id", "type": "TEXT", "primaryKey": True, "nullable": False},
                    {"name": "app_no", "type": "TEXT", "nullable": True}, {"name": "student_name", "type": "TEXT", "nullable": False},
                    {"name": "dob_age", "type": "TEXT", "nullable": True}, {"name": "course", "type": "TEXT", "nullable": True},
                    {"name": "department", "type": "TEXT", "nullable": True}, {"name": "degree_year", "type": "TEXT", "nullable": True},
                    {"name": "dept_number", "type": "TEXT", "nullable": True}, {"name": "mobile_no", "type": "TEXT", "nullable": True},
                    {"name": "father_name", "type": "TEXT", "nullable": True}, {"name": "father_occupation", "type": "TEXT", "nullable": True},
                    {"name": "forenoon_meal", "type": "BOOLEAN", "nullable": False}, {"name": "afternoon_meal", "type": "BOOLEAN", "nullable": False},
                    {"name": "annual_income", "type": "TEXT", "nullable": True}, {"name": "distance_km", "type": "TEXT", "nullable": True},
                    {"name": "permanent_address", "type": "TEXT", "nullable": True}, {"name": "permanent_pin", "type": "TEXT", "nullable": True},
                    {"name": "local_address", "type": "TEXT", "nullable": True}, {"name": "local_pin", "type": "TEXT", "nullable": True},
                    {"name": "landline", "type": "TEXT", "nullable": True}, {"name": "employment_type", "type": "TEXT", "nullable": True},
                    {"name": "last_year_id", "type": "TEXT", "nullable": True}, {"name": "student_photo_base64", "type": "TEXT", "nullable": True},
                    {"name": "applicant_signature_base64", "type": "TEXT", "nullable": True}, {"name": "income_proof_filename", "type": "TEXT", "nullable": True},
                    {"name": "status", "type": "TEXT", "nullable": False}, {"name": "submitted_at", "type": "DATE", "nullable": False}
                ],
                "rows": []
            },
            "import_logs": {
                "columns": [
                    {"name": "id", "type": "TEXT"}, {"name": "filename", "type": "TEXT"},
                    {"name": "records_imported", "type": "INTEGER"}, {"name": "status", "type": "TEXT"},
                    {"name": "created_at", "type": "DATETIME"}, {"name": "error", "type": "TEXT", "nullable": True}
                ],
                "rows": []
            },
            "export_logs": {
                "columns": [
                    {"name": "id", "type": "TEXT"}, {"name": "filename", "type": "TEXT"},
                    {"name": "records_exported", "type": "INTEGER"}, {"name": "format", "type": "TEXT"},
                    {"name": "created_at", "type": "DATETIME"}
                ],
                "rows": []
            },
            "audit_logs": {
                "columns": [
                    {"name": "id", "type": "TEXT"}, {"name": "username", "type": "TEXT"},
                    {"name": "action", "type": "TEXT"}, {"name": "table_name", "type": "TEXT"},
                    {"name": "details", "type": "TEXT"}, {"name": "created_at", "type": "DATETIME"}
                ],
                "rows": []
            },
            "system_roles": {
                "columns": [
                    {"name": "id", "type": "TEXT"}, {"name": "role_name", "type": "TEXT"},
                    {"name": "description", "type": "TEXT"}
                ],
                "rows": [
                    {"id": "ROLE_ADMIN", "role_name": "admin", "description": "System administrator with full access"},
                    {"id": "ROLE_APPROVAL", "role_name": "approval_staff", "description": "Staff who approve meal tokens"},
                    {"id": "ROLE_CANTEEN", "role_name": "canteen_staff", "description": "Canteen counter staff who verify tokens"},
                    {"id": "ROLE_STUDENT", "role_name": "student", "description": "Student with meal entitlement"}
                ]
            }
        }

        state = {
            "users": user_rows,
            "import_logs": [],
            "export_logs": [],
            "audit_logs": [],
            "meal_registrations": [],
            "tables": tables
        }
        cur.execute("INSERT INTO app_state (id, data) VALUES (1, %s)", (json.dumps(state, default=str, ensure_ascii=False),))

def _gen_qr_secret(sid):
    return hashlib.sha256(f"{sid}:{QR_HMAC_SECRET}:{uuid.uuid4().hex}".encode()).hexdigest()[:32]

def _log_audit(username, action, table_name, details):
    if isinstance(username, dict):
        username = username.get('username') or 'admin'
    if not username or not isinstance(username, str):
        username = 'admin'
    try:
        conn = get_db()
        try:
            with conn.cursor() as cur:
                aid = "aud_" + uuid.uuid4().hex[:9]
                cur.execute("INSERT INTO audit_logs (id,username,action,table_name,details) VALUES (%s,%s,%s,%s,%s)",
                            (aid, str(username), str(action), str(table_name), str(details)))
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        logger.warning("Audit log error: %s", e)

def _log_scan(scanner_id, scanner_role, scan_type, payload, student_id, token_uid, result, detail=None):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("""INSERT INTO scan_audit_log (scanner_id,scanner_role,scan_type,payload,student_id,token_uid,result,detail)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (scanner_id, scanner_role, scan_type, payload, student_id, token_uid, result, detail))
    finally:
        conn.close()

def _coerce_value(val, target_type):
    """Python translation of importer.ts coerceValue(). Coerce a raw value to the target column type."""
    if val is None or val == '':
        return None
    s = str(val).strip()
    if target_type == 'NUMBER':
        try:
            return float(s) if '.' in s else int(s)
        except (ValueError, TypeError):
            return None
    elif target_type == 'BOOLEAN':
        low = s.lower()
        if low in ('true', '1', 'yes'):
            return True
        if low in ('false', '0', 'no'):
            return False
        return bool(val)
    elif target_type == 'DATE':
        try:
            from dateutil.parser import parse as date_parse
            return date_parse(s).strftime('%Y-%m-%d')
        except Exception:
            return s
    else:  # TEXT
        return str(val)

def _hmac_sign(data):
    return hmac.new(QR_HMAC_SECRET.encode(), data.encode(), hashlib.sha256).hexdigest()[:16]

def _hmac_verify(data, signature):
    return hmac.compare_digest(_hmac_sign(data), signature)

def _generate_token_uid(meal_type):
    prefix = "FN" if meal_type == 'forenoon' else "AN"
    return f"{prefix}-{uuid.uuid4().hex[:8].upper()}"

def _generate_student_qr(student):
    payload = json.dumps({"sid": student['student_id'], "ts": int(time.time())}, separators=(',', ':'))
    sig = _hmac_sign(payload)
    full = base64.urlsafe_b64encode(f"{payload}.{sig}".encode()).decode()
    return full

def _generate_token_qr(token_uid, student_id, meal_type, expiry_time):
    expiry_str = expiry_time.isoformat() if expiry_time else (datetime.datetime.now() + datetime.timedelta(hours=1)).isoformat()
    payload = json.dumps({"tu": token_uid, "sid": student_id, "mt": meal_type, "exp": expiry_str}, separators=(',', ':'))
    sig = _hmac_sign(payload)
    full = base64.urlsafe_b64encode(f"{payload}.{sig}".encode()).decode()
    return full

def _qr_image(data):
    qr = qrcode.QRCode(box_size=8, border=1)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf)
    buf.seek(0)
    return buf

def _append_credentials_csv(row_dict, fieldnames):
    path = CREDENTIALS_CSV_PATH
    file_exists = os.path.isfile(path)
    with open(path, 'a', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerow(row_dict)

def _check_active_window(meal_type):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, start_time, end_time, 
                       expiry_minutes
                FROM meal_windows
                WHERE meal_type = %s AND is_active = 1
                  AND (day_of_week IS NULL OR day_of_week = DAYOFWEEK(CURDATE()) - 1)
                  AND start_time <= CURTIME() AND end_time >= CURTIME()
                LIMIT 1
            """, (meal_type,))
            return cur.fetchone()
    finally:
        conn.close()

def _get_meal_type_label(meal_type):
    return 'forenoon' if meal_type == 'forenoon' else 'afternoon'

def _lazy_expire_tokens(conn):
    try:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE meal_tokens SET status = 'expired'
                WHERE status IN ('token_issued','staff_verified','approved','active','awaiting_scan')
                  AND ((expiry_time IS NOT NULL AND expiry_time < NOW())
                    OR (expiry_time IS NULL AND TIMESTAMPDIFF(SECOND, created_at, NOW()) > 1800))
            """)
    except Exception as e:
        logger.warning(f"Error in _lazy_expire_tokens: {e}")

# --- AUTH HELPERS ---

def generate_token(user):
    payload = {
        "id": user["id"], "username": user["username"],
        "role": user["role"], "display_name": user.get("display_name"),
        "exp": datetime.datetime.utcnow() + datetime.timedelta(hours=24)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")

def decode_token(token):
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
    except Exception:
        return None

from functools import wraps

def _get_current_user():
    from flask import g
    return getattr(g, 'user', None) or {}

def authenticate(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({"error": "Unauthorized: No token provided"}), 401
        token = auth_header.split(' ')[1]
        payload = decode_token(token)
        if not payload:
            return jsonify({"error": "Unauthorized: Invalid token"}), 401
        from flask import g
        g.user = payload
        return f(*args, **kwargs)
    return decorated

def require_role(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user = _get_current_user()
            if not user or user.get('role') not in roles:
                return jsonify({"error": "Forbidden: insufficient permissions"}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator

# --- AUTH ROUTES ---

def _verify_user_password(plain_pw, stored_hash):
    if not plain_pw or not stored_hash:
        return False
    try:
        if stored_hash.startswith('$2a$') or stored_hash.startswith('$2b$') or stored_hash.startswith('$2y$'):
            if bcrypt.checkpw(plain_pw.encode('utf-8'), stored_hash.encode('utf-8')):
                return True
    except Exception:
        pass
    try:
        from werkzeug.security import check_password_hash
        if check_password_hash(stored_hash, plain_pw):
            return True
    except Exception:
        pass
    if plain_pw == stored_hash:
        return True
    return False

@admin_bp.route('/system/status', methods=['GET'])
@admin_bp.route('/api/system/status', methods=['GET'])
def system_status():
    students, today_tokens, today_redeemed, active_windows = 0, 0, 0, 0
    total_tables, total_records, total_imports, total_exports = 0, 0, 0, 0
    
    try:
        conn = get_db()
        if not conn:
            return jsonify({
                "status": "OFFLINE",
                "online": False,
                "connected": False,
                "databaseEngine": f"MySQL ({MYSQL_HOST}:{MYSQL_PORT}/{MYSQL_DATABASE})",
                "error": "Database connection unavailable"
            }), 200

        try:
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                try:
                    cur.execute("SELECT COUNT(*) as c FROM student_meals")
                    students = cur.fetchone().get('c', 0)
                except Exception: pass
                
                try:
                    cur.execute("SELECT COUNT(*) as c FROM meal_tokens WHERE DATE(created_at) = CURDATE()")
                    today_tokens = cur.fetchone().get('c', 0)
                except Exception: pass
                
                try:
                    cur.execute("SELECT COUNT(*) as c FROM meal_tokens WHERE DATE(created_at) = CURDATE() AND status = 'redeemed'")
                    today_redeemed = cur.fetchone().get('c', 0)
                except Exception: pass
                
                try:
                    cur.execute("SELECT COUNT(*) as c FROM meal_windows WHERE is_active = 1")
                    active_windows = cur.fetchone().get('c', 0)
                except Exception: pass
                
                try:
                    cur.execute("SELECT data FROM app_state WHERE id = 1")
                    state_row = cur.fetchone()
                    if state_row and state_row.get('data'):
                        state_data = json.loads(state_row['data'])
                        tables_dict = state_data.get('tables', {})
                        total_tables = len(tables_dict)
                        for tbl in tables_dict.values():
                            total_records += len(tbl.get('rows', []))
                except Exception: pass
                
                try:
                    cur.execute("SELECT COUNT(*) as c FROM import_logs")
                    total_imports = cur.fetchone().get('c', 0)
                except Exception: pass
                
                try:
                    cur.execute("SELECT COUNT(*) as c FROM export_logs")
                    total_exports = cur.fetchone().get('c', 0)
                except Exception: pass
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

        return jsonify({
            "status": "ONLINE",
            "online": True,
            "connected": True,
            "totalStudents": students,
            "tokensToday": today_tokens,
            "redeemedToday": today_redeemed,
            "activeWindows": active_windows,
            "totalTables": total_tables,
            "totalRecords": total_records,
            "totalImports": total_imports,
            "totalExports": total_exports,
            "diskUsage": "N/A",
            "databaseEngine": f"MySQL ({MYSQL_HOST}:{MYSQL_PORT}/{MYSQL_DATABASE})",
            "serverTime": datetime.datetime.now().isoformat()
        })
    except Exception as e:
        logger.error("System status probe failed: %s", e)
        return jsonify({
            "status": "OFFLINE",
            "online": False,
            "connected": False,
            "databaseEngine": f"MySQL ({MYSQL_HOST}:{MYSQL_PORT}/{MYSQL_DATABASE})",
            "error": str(e)
        }), 200

@admin_bp.route('/login', methods=['POST'])
@admin_bp.route('/auth/login', methods=['POST'])
@admin_bp.route('/api/auth/login', methods=['POST'])
@limiter.limit("60 per minute")
def auth_login():
    try:
        data = request.json or {}
        username = data.get('username', '').strip()
        password = data.get('password', '')
        if not username or not password:
            return jsonify({"error": "Username and password required"}), 400
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM users WHERE LOWER(username) = LOWER(%s)", (username,))
                user = cur.fetchone()
        finally:
            conn.close()
            
        if not user or not _verify_user_password(password, user.get('password_hash', '')):
            return jsonify({"error": "Invalid username or password"}), 401
        
        # Enforce administrative entry point isolation
        if user.get('role') == 'student':
            return jsonify({"error": "Access Denied: Student accounts must use the Student Portal link."}), 403

        token = generate_token(user)
        _log_audit(username, 'LOGIN', 'users', 'Logged in')
        return jsonify({
            "token": token,
            "user": {"id": user["id"], "username": user["username"], "email": user["email"],
                     "role": user["role"], "display_name": user.get("display_name")}
        })
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/auth/check-user', methods=['POST'])
@admin_bp.route('/api/auth/check-user', methods=['POST'])
def staff_check_user():
    try:
        data = request.get_json(silent=True) or {}
        username = (data.get('username') or data.get('register_no') or data.get('staff_id') or '').strip()
        if not username:
            return jsonify({"error": "Username / Staff ID is required"}), 400
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT id, username, email, display_name, role FROM users WHERE LOWER(TRIM(username)) = LOWER(TRIM(%s)) OR LOWER(TRIM(id)) = LOWER(TRIM(%s)) LIMIT 1", (username, username))
                user = cur.fetchone()
        finally:
            conn.close()
            
        if not user:
            return jsonify({"exists": False, "error": "Staff user with this Username / ID was not found in the database."}), 404
            
        return jsonify({
            "exists": True,
            "username": user['username'],
            "display_name": user.get('display_name') or user['username'],
            "role": user.get('role'),
            "has_email": bool(user.get('email'))
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@admin_bp.route('/auth/forgot-password', methods=['POST'])
@admin_bp.route('/api/auth/forgot-password', methods=['POST'])
@admin_bp.route('/auth/staff-forgot-password', methods=['POST'])
def staff_forgot_password():
    try:
        data = request.get_json(silent=True) or {}
        username = (data.get('username') or data.get('register_no') or data.get('staff_id') or '').strip()
        provided_email = (data.get('email') or '').strip()
        
        if not username:
            return jsonify({"error": "Please enter your Staff Username / ID"}), 400
        if not provided_email:
            return jsonify({"error": "Please enter your registered Email Address"}), 400
            
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM users WHERE LOWER(TRIM(username)) = LOWER(TRIM(%s)) OR LOWER(TRIM(id)) = LOWER(TRIM(%s)) LIMIT 1", (username, username))
                user = cur.fetchone()
                
                if not user:
                    return jsonify({"error": "Staff user with this Username / ID was not found in the database."}), 404
                    
                db_email = (user.get('email') or '').strip()
                if not db_email or db_email.lower() != provided_email.lower():
                    return jsonify({"error": "The entered email address does not match our records for this staff account."}), 400
                    
                # Generate clean unambiguous random password
                import secrets
                chars_upper = "ABCDEFGHJKLMNPQRSTUVWXYZ"
                chars_lower = "abcdefghijkmnopqrstuvwxyz"
                chars_digits = "23456789"
                chars_symbols = "@#!"

                raw_pwd = [
                    secrets.choice(chars_upper),
                    secrets.choice(chars_lower),
                    secrets.choice(chars_digits),
                    secrets.choice(chars_symbols)
                ]
                all_chars = chars_upper + chars_lower + chars_digits + chars_symbols
                raw_pwd += [secrets.choice(all_chars) for _ in range(6)]
                secrets.SystemRandom().shuffle(raw_pwd)
                new_password = "".join(raw_pwd)
                
                import bcrypt
                salt = bcrypt.gensalt()
                hashed = bcrypt.hashpw(new_password.encode('utf-8'), salt).decode('utf-8')
                cur.execute("UPDATE users SET password_hash = %s WHERE id = %s", (hashed, user['id']))
                conn.commit()
        finally:
            conn.close()

        # Build and dispatch email
        from admin_backend.email_service import get_smtp_config, _send_mime_message
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        cfg = get_smtp_config()
        msg = MIMEMultipart('alternative')
        msg['From'] = cfg['from']
        msg['To'] = provided_email
        msg['Subject'] = 'Ramakrishna Mission Vidyapith - Staff Portal Password Reset'

        text_body = f"""Dear {user.get('display_name') or user.get('username')},

Your password for the RKMVC Staff / Admin Portal has been reset successfully.

Username / Staff ID: {user['username']}
New Password: {new_password}

Login URL: http://localhost:5050/admin-login/

Please keep this password secure. You can now login with your new password.

Ramakrishna Mission Vidyapith
Mylapore, Chennai - 600 004."""

        html_body = f"""<html><body style="font-family: Arial, sans-serif; color: #333; margin: 0; padding: 10px;">
<div style="max-width: 600px; margin: 0 auto; background: #fff8f0; border: 1px solid #fbd5a5; border-radius: 12px; overflow: hidden;">
<div style="background: #ea580c; padding: 20px; text-align: center;">
<h2 style="color: #fff; margin: 0; font-size: 18px; font-weight: bold;">Ramakrishna Mission Vidyapith</h2>
<p style="color: #ffedd5; margin: 4px 0 0; font-size: 12px;">Staff Portal Password Reset</p>
</div>
<div style="padding: 24px;">
<p style="font-size: 14px;">Dear <strong>{user.get('display_name') or user.get('username')}</strong>,</p>
<p style="font-size: 13px;">Your password for the Staff Portal has been successfully reset.</p>
<div style="background: #fff; border: 1px solid #fed7aa; border-radius: 10px; padding: 18px; margin: 18px 0; text-align: center;">
<p style="font-size: 11px; color: #9a3412; font-weight: bold; text-transform: uppercase; margin: 0 0 6px;">Your New Unique Password</p>
<p style="font-size: 20px; font-family: monospace; font-weight: bold; color: #7c2d12; background: #fff7ed; padding: 10px 16px; border: 1px solid #fbd5a5; border-radius: 8px; margin: 0; display: inline-block;">{new_password}</p>
</div>
<p style="font-size: 13px; color: #4b5563;">You can use this password to sign into the staff portal.</p>
</div>
<div style="background: #ffedd5; padding: 12px; text-align: center; font-size: 11px; color: #9a3412;">
Ramakrishna Mission Vidyapith &bull; Mylapore, Chennai - 600 004.
</div>
</div></body></html>"""

        msg.attach(MIMEText(text_body, 'plain'))
        msg.attach(MIMEText(html_body, 'html'))

        try:
            _send_mime_message(cfg, msg)
        except Exception as mail_err:
            print("Notice: Staff password reset email alert:", mail_err, flush=True)

        _log_audit(username, 'FORGOT_PASSWORD', 'users', 'Staff password reset requested')
        return jsonify({
            "message": "Password reset successfully. A new password has been sent to your email address.",
            "success": True
        }), 200
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/auth/me', methods=['GET'])
@admin_bp.route('/api/auth/me', methods=['GET'])
@admin_bp.route('/me', methods=['GET'])
def auth_me():
    try:
        auth_header = request.headers.get('Authorization', '')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({"error": "Unauthorized: No token provided"}), 401
        token = auth_header.split(' ')[1]
        user_payload = decode_token(token) or {}
        user_id = user_payload.get('id', 'usr_admin')
        username = user_payload.get('username', 'admin')
        
        user = None
        try:
            conn = get_db()
            try:
                with conn.cursor() as cur:
                    cur.execute("SELECT id,username,email,role,display_name FROM users WHERE id = %s OR username = %s", (user_id, username))
                    user = cur.fetchone()
            finally:
                conn.close()
        except Exception:
            user = None

        if not user:
            user = {
                "id": user_id,
                "username": username,
                "email": f"{username}@example.com",
                "role": user_payload.get('role', 'admin'),
                "display_name": user_payload.get('display_name', 'System Administrator')
            }
        return jsonify({"user": user})
    except Exception as e:
        return jsonify({"user": {
            "id": "usr_admin",
            "username": "admin",
            "email": "admin@example.com",
            "role": "admin",
            "display_name": "System Administrator"
        }})

@admin_bp.route('/auth/register', methods=['POST'])
@admin_bp.route('/api/auth/register', methods=['POST'])
@authenticate
@require_role('admin')
def auth_register():
    try:
        data = request.get_json(silent=True) or {}
        username, email, password = data.get('username'), data.get('email'), data.get('password')
        if not username or not email or not password:
            return jsonify({"error": "Username, email, and password required"}), 400
        role = data.get('role', 'approval_staff')
        display_name = data.get('display_name', username)
        # Admins may only provision Staff or Admin accounts here.
        # Student accounts are created exclusively via the registration flow.
        allowed_roles = ('admin', 'approval_staff', 'canteen_staff')
        if role not in allowed_roles:
            return jsonify({"error": f"Admins can only create Staff or Admin accounts. Allowed roles: {', '.join(allowed_roles)}"}), 403
        # Validate field lengths up front so we return a clear 400 instead of
        # letting MySQL reject an oversized value and surfacing a generic 500.
        field_limits = {"username": 50, "email": 100, "display_name": 100}
        for field_name, value in (("username", username), ("email", email), ("display_name", display_name)):
            if value and len(value) > field_limits[field_name]:
                return jsonify({
                    "error": f"{field_name.replace('_', ' ').title()} must be {field_limits[field_name]} characters or fewer (got {len(value)})."
                }), 400
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM users WHERE username = %s OR email = %s", (username, email))
                if cur.fetchone():
                    return jsonify({"error": "Username or email already taken"}), 400
                uid = "usr_" + uuid.uuid4().hex[:9]
                pw_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode()
                cur.execute("INSERT INTO users (id,username,email,password_hash,role,display_name) VALUES (%s,%s,%s,%s,%s,%s)",
                            (uid, username, email, pw_hash, role, display_name))
                conn.commit()
        finally:
            conn.close()
        _log_audit(_get_auditor_username(), 'REGISTER', 'users', f"Created user {username} with role {role}")
        return jsonify({"message": "User created", "id": uid}), 201
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/auth/users', methods=['GET'])
@admin_bp.route('/api/auth/users', methods=['GET'])
@authenticate
@require_role('admin')
def auth_list_users():
    try:
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id,username,email,role,display_name,created_at FROM users "
                    "WHERE role IN ('admin','approval_staff','canteen_staff') ORDER BY created_at DESC"
                )
                users = cur.fetchall()
        finally:
            conn.close()
        return jsonify({"users": _sanitize_for_json(users)})
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

# --- MEAL TIMINGS MANAGEMENT (APP STATE) ---

@admin_bp.route('/meal-timings', methods=['GET'])
@admin_bp.route('/admin/meal-timings', methods=['GET'])
@admin_bp.route('/api/admin/meal-timings', methods=['GET'])
@authenticate
@require_role('admin', 'approval_staff', 'canteen_staff')
def get_meal_timings():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT data FROM app_state WHERE id = 1")
            row = cur.fetchone()
            if row and row.get('data'):
                try:
                    data = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
                    return jsonify(data.get('meal_timings', {
                        "forenoon": {"start": "07:30", "end": "10:00"},
                        "afternoon": {"start": "12:00", "end": "14:30"}
                    }))
                except Exception:
                    pass
            return jsonify({
                "forenoon": {"start": "07:30", "end": "10:00"},
                "afternoon": {"start": "12:00", "end": "14:30"}
            })
    finally:
        conn.close()

@admin_bp.route('/meal-timings', methods=['POST', 'PUT'])
@admin_bp.route('/admin/meal-timings', methods=['POST', 'PUT'])
@admin_bp.route('/api/admin/meal-timings', methods=['POST', 'PUT'])
@authenticate
@require_role('admin')
def update_meal_timings():
    data = request.json or {}
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT data FROM app_state WHERE id = 1")
            row = cur.fetchone()
            state_data = {}
            if row and row.get('data'):
                state_data = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
            
            state_data['meal_timings'] = data
            cur.execute("""
                INSERT INTO app_state (id, data) VALUES (1, %s)
                ON DUPLICATE KEY UPDATE data = VALUES(data)
            """, (json.dumps(state_data),))
            conn.commit()
            return jsonify({"message": "Meal timings updated successfully", "meal_timings": data})
    finally:
        conn.close()

# --- STUDENT MEALS API ---

@admin_bp.route('/students', methods=['GET'])
@admin_bp.route('/api/students', methods=['GET'])
@authenticate
@require_role('admin', 'approval_staff', 'canteen_staff')
def get_students():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM student_meals ORDER BY student_id ASC")
            rows = cur.fetchall()
            for r in rows:
                r['degree_year'] = _format_degree_year(r.get('degree_year'))
            return jsonify(rows)
    finally:
        conn.close()

@admin_bp.route('/students/<student_id>', methods=['GET'])
@admin_bp.route('/api/students/<student_id>', methods=['GET'])
@authenticate
def get_student(student_id):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM student_meals WHERE student_id = %s", (student_id,))
            s = cur.fetchone()
            if not s:
                return jsonify({"error": "Student not found"}), 404
            return jsonify(s)
    finally:
        conn.close()

# --- STUDENT QR ENDPOINT ---

@admin_bp.route('/qr', methods=['GET'])
@admin_bp.route('/student/qr', methods=['GET'])
@admin_bp.route('/api/student/qr', methods=['GET'])
@authenticate
@require_role('student')
def get_student_qr():
    student_id = _get_current_user().get('student_id')
    if not student_id:
        return jsonify({"error": "No student profile linked to this account"}), 400
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM student_meals WHERE student_id = %s", (student_id,))
            student = cur.fetchone()
            if not student:
                return jsonify({"error": "Student not found"}), 404
    finally:
        conn.close()
    qr_data = _generate_student_qr(student)
    return jsonify({"qr_data": qr_data, "student_id": student_id, "name": student['name']})

@admin_bp.route('/qr-image', methods=['GET'])
@admin_bp.route('/student/qr-image', methods=['GET'])
@admin_bp.route('/api/student/qr-image', methods=['GET'])
@authenticate
@require_role('student')
def get_student_qr_image():
    student_id = _get_current_user().get('student_id')
    if not student_id:
        return jsonify({"error": "No student profile linked"}), 400
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM student_meals WHERE student_id = %s", (student_id,))
            student = cur.fetchone()
    finally:
        conn.close()
    if not student:
        return jsonify({"error": "Student not found"}), 404
    qr_data = _generate_student_qr(student)
    buf = _qr_image(qr_data)
    return send_file(buf, mimetype='image/png')

# --- MEAL WINDOWS API ---

# --- MEAL CONFIG (single config replacing individual day/window CRUD) ---

WEEKDAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

def _format_time_str(val, default="07:30"):
    if val is None or val == '':
        return default
    if isinstance(val, datetime.timedelta):
        total = int(val.total_seconds())
        h = (total // 3600) % 24
        m = (total % 3600) // 60
        return f"{h:02d}:{m:02d}"
    if isinstance(val, (datetime.time, datetime.datetime)):
        return val.strftime("%H:%M")
    s = str(val).strip()
    parts = s.split(':')
    if len(parts) >= 2:
        try:
            h = int(parts[0])
            m = int(parts[1])
            return f"{h:02d}:{m:02d}"
        except ValueError:
            pass
    return default

def _add_one_year(dt):
    try:
        return dt.replace(year=dt.year + 1)
    except ValueError:
        # Handling Feb 29 leap year to non-leap year (Feb 28)
        return dt.replace(year=dt.year + 1, day=28)

def _execute_year_migration(conn):
    """Batch migrates student academic years: 1st Year -> 2nd Year -> 3rd Year -> Graduated."""
    with conn.cursor() as cur:
        # 3rd Year -> Graduated (disable active meal eligibility for graduated students in student_meals)
        cur.execute("""
            UPDATE student_meals 
            SET degree_year = 'Graduated', forenoon_meal = 0, afternoon_meal = 0 
            WHERE degree_year LIKE '3rd%' OR degree_year = '3';
        """)
        
        # 2nd Year -> 3rd Year
        cur.execute("UPDATE student_meals SET degree_year = '3rd Year' WHERE degree_year LIKE '2nd%' OR degree_year = '2';")
        
        # 1st Year -> 2nd Year
        cur.execute("UPDATE student_meals SET degree_year = '2nd Year' WHERE degree_year LIKE '1st%' OR degree_year = '1';")
        
        conn.commit()

def _check_and_run_automated_year_migration(conn):
    """Checks if year_migration_date in app_state is set and <= NOW().
       If so, executes year migration and advances scheduled datetime by +1 year until > NOW()."""
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT data FROM app_state WHERE id = 1 FOR UPDATE")
            row = cur.fetchone()
            if not row or not row.get('data'):
                return
            state = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
            meal_cfg = state.get('meal_config')
            if not meal_cfg or not isinstance(meal_cfg, dict):
                return
            
            migration_str = meal_cfg.get('year_migration_date')
            if not migration_str or not str(migration_str).strip():
                return
            
            migration_str = str(migration_str).strip()
            dt_obj = None
            formats = (
                "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                "%d-%m-%YT%H:%M:%S", "%d-%m-%YT%H:%M", "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M",
                "%d/%m/%YT%H:%M:%S", "%d/%m/%YT%H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                "%m/%d/%YT%H:%M:%S", "%m/%d/%YT%H:%M", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
                "%Y/%m/%YT%H:%M:%S", "%Y/%m/%YT%H:%M", "%Y/%m/%Y %H:%M:%S", "%Y/%m/%Y %H:%M"
            )
            for fmt in formats:
                try:
                    dt_obj = datetime.datetime.strptime(migration_str, fmt)
                    break
                except ValueError:
                    continue
            
            if not dt_obj:
                logger.warning(f"Could not parse scheduled migration date: {migration_str}")
                return

            now = datetime.datetime.now()
            if now >= dt_obj:
                logger.info(f"Automated year migration triggered for scheduled date {migration_str} (current time {now})")
                # 1. Execute year migration
                _execute_year_migration(conn)
                
                # 2. Advance next migration date by +1 year until it is strictly in the future
                next_dt = dt_obj
                while next_dt <= now:
                    next_dt = _add_one_year(next_dt)
                
                new_migration_str = next_dt.strftime("%Y-%m-%dT%H:%M")
                meal_cfg['year_migration_date'] = new_migration_str
                state['meal_config'] = meal_cfg
                
                cur.execute("UPDATE app_state SET data = %s WHERE id = 1", (json.dumps(state, ensure_ascii=False),))
                conn.commit()
                _log_audit('SYSTEM', 'AUTOMATED_YEAR_MIGRATION', 'student_meals', f"Automated year migration executed. Next migration set to {new_migration_str}")
    except Exception as e:
        logger.error("Error in _check_and_run_automated_year_migration: %s", e)

def _get_meal_config(conn):
    try:
        _check_and_run_automated_year_migration(conn)
    except Exception as e:
        logger.error("Error running year migration check in _get_meal_config: %s", e)

    year_migration_date = ""
    with conn.cursor() as cur:
        cur.execute("SELECT data FROM app_state WHERE id = 1")
        row = cur.fetchone()
        if row and row.get('data'):
            state = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
            cfg_state = state.get('meal_config') or state.get('meal_timings')
            if cfg_state and isinstance(cfg_state, dict):
                year_migration_date = str(cfg_state.get('year_migration_date') or '').strip()

        # First query meal_windows SQL table if configured
        try:
            cur.execute("""
                SELECT meal_type, MIN(start_time) as start_t, MAX(end_time) as end_t, 
                       MAX(expiry_minutes) as expiry_m 
                FROM meal_windows WHERE is_active = 1 GROUP BY meal_type
            """)
            rows = cur.fetchall()
            if rows:
                res = {}
                for r in rows:
                    mt = r['meal_type']
                    def_start = "07:30" if mt == 'forenoon' else "11:30"
                    def_end = "10:00" if mt == 'forenoon' else "19:30"
                    st = _format_time_str(r['start_t'], def_start)
                    et = _format_time_str(r['end_t'], def_end)
                    exp = int(r['expiry_m']) if r.get('expiry_m') is not None else 30
                    if exp == 15:
                        exp = 30
                    res[mt] = {"start": st, "end": et, "expiry": exp}
                
                cur.execute("SELECT DISTINCT day_of_week FROM meal_windows WHERE is_active = 1")
                day_rows = cur.fetchall()
                res['days'] = [d['day_of_week'] for d in day_rows if d.get('day_of_week') is not None]
                if 'forenoon' in res or 'afternoon' in res:
                    if 'forenoon' not in res:
                        res['forenoon'] = {"start": "07:30", "end": "10:00", "expiry": 30}
                    if 'afternoon' not in res:
                        res['afternoon'] = {"start": "11:30", "end": "19:30", "expiry": 30}
                    res['year_migration_date'] = year_migration_date
                    return res
        except Exception:
            pass

        # Fallback to app_state JSON
        if row and row.get('data'):
            state = json.loads(row['data']) if isinstance(row['data'], str) else row['data']
            cfg = state.get('meal_config') or state.get('meal_timings')
            if cfg and isinstance(cfg, dict):
                if 'forenoon' in cfg:
                    cfg['forenoon']['start'] = _format_time_str(cfg['forenoon'].get('start'), "07:30")
                    cfg['forenoon']['end'] = _format_time_str(cfg['forenoon'].get('end'), "10:00")
                    if cfg['forenoon'].get('expiry') in (15, None):
                        cfg['forenoon']['expiry'] = 30
                if 'afternoon' in cfg:
                    cfg['afternoon']['start'] = _format_time_str(cfg['afternoon'].get('start'), "11:30")
                    cfg['afternoon']['end'] = _format_time_str(cfg['afternoon'].get('end'), "19:30")
                    if cfg['afternoon'].get('expiry') in (15, None):
                        cfg['afternoon']['expiry'] = 30
                cfg['year_migration_date'] = year_migration_date
                return cfg
        def_cfg = _default_meal_config()
        def_cfg['year_migration_date'] = year_migration_date
        return def_cfg

def _default_meal_config():
    return {
        "days": [0, 1, 2, 3, 4, 5, 6],
        "forenoon": {"start": "07:30", "end": "10:00", "expiry": 30},
        "afternoon": {"start": "12:00", "end": "14:30", "expiry": 30},
        "year_migration_date": ""
    }

def _sync_meal_windows(conn, cfg):
    with conn.cursor() as cur:
        cur.execute("DELETE FROM meal_windows")
        days = cfg['days']
        for meal_type in ('forenoon', 'afternoon'):
            m = cfg[meal_type]
            start_t = m['start']
            end_t = m['end']
            expiry = int(m['expiry'])
            for dow in days:
                cur.execute(
                    "INSERT INTO meal_windows (meal_type,day_of_week,start_time,end_time,expiry_minutes,is_active) VALUES (%s,%s,%s,%s,%s,1)",
                    (meal_type, dow, start_t, end_t, expiry)
                )

@admin_bp.route('/meal-config', methods=['GET'])
@admin_bp.route('/api/meal-config', methods=['GET'])
@admin_bp.route('/api/public/meal-config', methods=['GET'])
def get_meal_config():
    conn = get_db()
    try:
        cfg = _get_meal_config(conn)
        return jsonify(cfg)
    finally:
        conn.close()

@admin_bp.route('/meal-config', methods=['PUT'])
@admin_bp.route('/api/meal-config', methods=['PUT'])
@authenticate
@require_role('admin')
def update_meal_config():
    try:
        data = request.json or {}
        days = data.get('days')
        if days is None or not isinstance(days, list) or not all(isinstance(d, int) for d in days):
            return jsonify({"error": "days must be a list of integers (0=Mon..6=Sun)"}), 400
        for meal_type in ('forenoon', 'afternoon'):
            m = data.get(meal_type)
            if not isinstance(m, dict) or not str(m.get('start', '')).strip() or not str(m.get('end', '')).strip() or 'expiry' not in m:
                return jsonify({"error": f"Please enter valid start and end times for {meal_type.capitalize()} (e.g., 07:30)."}), 400
        
        fn_exp = int(data['forenoon']['expiry'])
        an_exp = int(data['afternoon']['expiry'])
        raw_mig_date = str(data.get('year_migration_date') or '').strip()
        normalized_mig_date = raw_mig_date
        if raw_mig_date:
            dt_parse = None
            formats = (
                "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                "%d-%m-%YT%H:%M:%S", "%d-%m-%YT%H:%M", "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M",
                "%d/%m/%YT%H:%M:%S", "%d/%m/%YT%H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                "%m/%d/%YT%H:%M:%S", "%m/%d/%YT%H:%M", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M"
            )
            for fmt in formats:
                try:
                    dt_parse = datetime.datetime.strptime(raw_mig_date, fmt)
                    break
                except ValueError:
                    continue
            if dt_parse:
                normalized_mig_date = dt_parse.strftime("%Y-%m-%dT%H:%M")

        cfg = {
            "days": days,
            "forenoon": {"start": str(data['forenoon']['start']).strip(), "end": str(data['forenoon']['end']).strip(), "expiry": fn_exp},
            "afternoon": {"start": str(data['afternoon']['start']).strip(), "end": str(data['afternoon']['end']).strip(), "expiry": an_exp},
            "year_migration_date": normalized_mig_date
        }
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT data FROM app_state WHERE id = 1 FOR UPDATE")
                row = cur.fetchone()
                conn.autocommit(False)
                if row:
                    state = json.loads(row['data'])
                else:
                    state = {}
                state['meal_config'] = cfg
                if row:
                    cur.execute("UPDATE app_state SET data = %s WHERE id = 1", (json.dumps(state, ensure_ascii=False),))
                else:
                    cur.execute("INSERT INTO app_state (id, data) VALUES (1, %s)", (json.dumps(state, ensure_ascii=False),))
                _sync_meal_windows(conn, cfg)
                conn.commit()
            
            # Immediately run migration check if the date was set in the past or is due
            _check_and_run_automated_year_migration(conn)
            final_cfg = _get_meal_config(conn)
        except Exception as sql_err:
            conn.rollback()
            logger.error("Error updating meal_config: %s", sql_err)
            return jsonify({"error": f"Failed to save meal windows: {str(sql_err)}"}), 400
        finally:
            conn.close()
        _log_audit(_get_auditor_username(), 'UPDATE_CONFIG', 'meal_config', f"Updated meal config: {len(days)} days, forenoon {cfg['forenoon']['start']}-{cfg['forenoon']['end']}, afternoon {cfg['afternoon']['start']}-{cfg['afternoon']['end']}, year_migration_date {year_migration_date}")
        return jsonify(final_cfg)
    except Exception as e:
        logger.error("Error in update_meal_config: %s", e)
        return jsonify({"error": str(e) if str(e) else sani(e)}), 400

def _start_migration_scheduler():
    def _scheduler_loop():
        time.sleep(5)
        while True:
            try:
                conn = get_db()
                try:
                    _check_and_run_automated_year_migration(conn)
                finally:
                    conn.close()
            except Exception:
                pass
            time.sleep(30)
    
    t = threading.Thread(target=_scheduler_loop, daemon=True, name="YearMigrationSchedulerThread")
    t.start()

try:
    _start_migration_scheduler()
except Exception as _e:
    logger.warning("Scheduler startup failed: %s", _e)

# Keep the /active endpoint for blueprints
@admin_bp.route('/meal-windows/active', methods=['GET'])
@admin_bp.route('/api/meal-windows/active', methods=['GET'])
@authenticate
def get_active_windows():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT * FROM meal_windows
                WHERE is_active = 1
                  AND (day_of_week IS NULL OR day_of_week = DAYOFWEEK(CURDATE()) - 1)
                  AND start_time <= CURTIME() AND end_time >= CURTIME()
                ORDER BY meal_type
            """)
            return jsonify(serialize_row(cur.fetchall()))
    finally:
        conn.close()

# --- TOKEN LIFECYCLE API ---

@admin_bp.route('/tokens/scan-student', methods=['POST'])
@admin_bp.route('/api/tokens/scan-student', methods=['POST'])
@authenticate
@require_role('approval_staff')
def scan_student_qr():
    try:
        data = request.json or {}
        scanned_payload = data.get('scanned_payload', '')
        scanner_id = _get_auditor_username()
        if not scanned_payload:
            return jsonify({"error": "No QR payload"}), 400
        try:
            decoded = base64.urlsafe_b64decode(scanned_payload.encode()).decode()
            payload_part, sig = decoded.rsplit('.', 1)
            qr_data = json.loads(payload_part)
        except Exception:
            _log_scan(scanner_id, 'approval_staff', 'student_id_qr', scanned_payload, None, None, 'invalid_signature')
            return jsonify({"status": "INVALID", "error": "Invalid QR format"}), 400
        if not _hmac_verify(payload_part, sig):
            _log_scan(scanner_id, 'approval_staff', 'student_id_qr', scanned_payload, None, None, 'invalid_signature')
            return jsonify({"status": "INVALID", "error": "QR signature invalid"}), 400
        student_id = qr_data.get('sid')
        if not student_id:
            return jsonify({"error": "Invalid QR data"}), 400
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM student_meals WHERE student_id = %s", (student_id,))
                student = cur.fetchone()
        finally:
            conn.close()
        if not student:
            _log_scan(scanner_id, 'approval_staff', 'student_id_qr', scanned_payload, student_id, None, 'not_found')
            return jsonify({"status": "INVALID", "error": "Student not found"}), 404
        active_window = _check_active_window('forenoon') or _check_active_window('afternoon')
        if not active_window:
            _log_scan(scanner_id, 'approval_staff', 'student_id_qr', scanned_payload, student_id, None, 'out_of_window')
            return jsonify({"status": "OUT_OF_WINDOW", "error": "No active meal window currently", "student": student}), 200
        meal_type = active_window['meal_type']
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, status FROM meal_tokens
                WHERE student_id = %s AND meal_type = %s
                  AND DATE(created_at) = CURDATE() AND status NOT IN ('rejected','expired')
                LIMIT 1
            """, (student_id, meal_type))
            existing = cur.fetchone()
        if existing:
            if existing['status'] in ('token_issued', 'staff_verified', 'approved'):
                _log_scan(scanner_id, 'approval_staff', 'student_id_qr', scanned_payload, student_id, None, 'duplicate_meal')
                return jsonify({"status": "DUPLICATE", "error": "Token already exists for this meal today", "student": student}), 200
        token_uid = _generate_token_uid(meal_type)
        now_dt = datetime.datetime.now()
        exp_mins = max(30, int(active_window.get('expiry_minutes') or 30))
        token_expiry = now_dt + datetime.timedelta(minutes=exp_mins)
        try:
            w_str = str(active_window['end_time'])
            window_end = datetime.datetime.strptime(w_str, '%H:%M:%S' if len(w_str) == 8 else '%H:%M').time()
            window_expiry = datetime.datetime.combine(now_dt.date(), window_end) + datetime.timedelta(minutes=exp_mins)
            expiry_dt = max(token_expiry, window_expiry)
        except Exception:
            expiry_dt = token_expiry
        conn2 = get_db()
        try:
            with conn2.cursor() as cur:
                cur.execute("""INSERT INTO meal_tokens (token_uid,student_id,meal_type,status,scanned_by,scanned_at,expiry_time)
                               VALUES (%s,%s,%s,'staff_verified',%s,NOW(),%s)""",
                            (token_uid, student_id, meal_type, scanner_id, expiry_dt))
                cur.execute("SELECT * FROM meal_tokens WHERE token_uid = %s", (token_uid,))
                token = cur.fetchone()
        finally:
            conn2.close()
        _log_scan(scanner_id, 'approval_staff', 'student_id_qr', scanned_payload, student_id, token_uid, 'success')
        return jsonify({
            "status": "VERIFIED", "token": token, "student": student,
            "active_window": {"meal_type": meal_type, "start": str(active_window['start_time']), "end": str(active_window['end_time'])}
        })
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/tokens/<token_uid>/approve', methods=['POST'])
@admin_bp.route('/api/tokens/<token_uid>/approve', methods=['POST'])
@authenticate
@require_role('approval_staff')
def approve_token(token_uid):
    try:
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM meal_tokens WHERE token_uid = %s", (token_uid,))
                token = cur.fetchone()
        finally:
            conn.close()
        if not token:
            return jsonify({"error": "Token not found"}), 404
        if token['status'] not in ('staff_verified',):
            return jsonify({"error": f"Cannot approve token in status '{token['status']}'"}), 400
        token_qr_data = _generate_token_qr(token_uid, token['student_id'], token['meal_type'], token['expiry_time'])
        conn2 = get_db()
        try:
            with conn2.cursor() as cur:
                cur.execute("""UPDATE meal_tokens SET status='approved', approved_by=%s, approved_at=NOW(),
                               token_qr_data=%s, token_issued_at=NOW() WHERE token_uid=%s""",
                            (_get_auditor_username(), token_qr_data, token_uid))
                cur.execute("SELECT * FROM meal_tokens WHERE token_uid = %s", (token_uid,))
                token = cur.fetchone()
                cur.execute("SELECT * FROM student_meals WHERE student_id = %s", (token['student_id'],))
                student = cur.fetchone()
        finally:
            conn2.close()
        _log_audit(_get_auditor_username(), 'APPROVE_TOKEN', 'meal_tokens', f"Approved token {token_uid} for {token['student_id']}")
        return jsonify({"status": "APPROVED", "token": token, "qr_data": token_qr_data, "student": student})
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/tokens/<token_uid>/reject', methods=['POST'])
@admin_bp.route('/api/tokens/<token_uid>/reject', methods=['POST'])
@authenticate
@require_role('approval_staff')
def reject_token(token_uid):
    try:
        data = request.json or {}
        reason = data.get('reason', 'Rejected by staff')
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("""UPDATE meal_tokens SET status='rejected', approved_by=%s,
                               reject_reason=%s WHERE token_uid=%s AND status NOT IN ('redeemed','expired','rejected')""",
                            (_get_auditor_username(), reason, token_uid))
                cur.execute("SELECT * FROM meal_tokens WHERE token_uid = %s", (token_uid,))
                token = cur.fetchone()
        finally:
            conn.close()
        if not token:
            return jsonify({"error": "Token not found"}), 404
        _log_audit(_get_auditor_username(), 'REJECT_TOKEN', 'meal_tokens', f"Rejected token {token_uid}: {reason}")
        return jsonify({"status": "REJECTED", "token": token})
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/tokens/redeem', methods=['POST'])
@admin_bp.route('/api/tokens/redeem', methods=['POST'])
@authenticate
@require_role('canteen_staff')
def redeem_token():
    try:
        data = request.json or {}
        scanned_payload = data.get('scanned_payload', '')
        scanner_id = _get_auditor_username()
        if not scanned_payload:
            return jsonify({"error": "No QR payload"}), 400
        try:
            decoded = base64.urlsafe_b64decode(scanned_payload.encode()).decode()
            payload_part, sig = decoded.rsplit('.', 1)
            qr_data = json.loads(payload_part)
        except Exception:
            _log_scan(scanner_id, 'canteen_staff', 'token_qr', scanned_payload, None, None, 'invalid_signature')
            return jsonify({"status": "INVALID", "error": "Invalid QR format"}), 400
        if not _hmac_verify(payload_part, sig):
            _log_scan(scanner_id, 'canteen_staff', 'token_qr', scanned_payload, None, None, 'invalid_signature')
            return jsonify({"status": "INVALID", "error": "QR signature invalid"}), 400
        token_uid = qr_data.get('tu') or qr_data.get('t')
        student_id = qr_data.get('sid') or qr_data.get('s')
        if not token_uid or not student_id:
            return jsonify({"error": "Invalid QR data"}), 400
        conn = get_db()
        try:
            _lazy_expire_tokens(conn)
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM meal_tokens WHERE token_uid = %s", (token_uid,))
                token = cur.fetchone()
                if not token:
                    _log_scan(scanner_id, 'canteen_staff', 'token_qr', scanned_payload, student_id, token_uid, 'not_found')
                    return jsonify({"status": "INVALID", "error": "Token not found in system"}), 404
                if token['status'] == 'redeemed':
                    _log_scan(scanner_id, 'canteen_staff', 'token_qr', scanned_payload, student_id, token_uid, 'already_redeemed')
                    return jsonify({"status": "ALREADY_REDEEMED", "error": "Token already redeemed",
                                    "redeemed_at": str(token['redeemed_at'])}), 200
                if token['status'] == 'expired':
                    _log_scan(scanner_id, 'canteen_staff', 'token_qr', scanned_payload, student_id, token_uid, 'expired')
                    return jsonify({"status": "EXPIRED", "error": "Token has expired"}), 200
                if token['status'] == 'rejected':
                    _log_scan(scanner_id, 'canteen_staff', 'token_qr', scanned_payload, student_id, token_uid, 'invalid_token')
                    return jsonify({"status": "INVALID", "error": f"Token was rejected: {token.get('reject_reason', 'No reason')}"}), 200
                if token['status'] not in ('token_issued', 'approved', 'staff_verified'):
                    _log_scan(scanner_id, 'canteen_staff', 'token_qr', scanned_payload, student_id, token_uid, 'invalid_token')
                    return jsonify({"status": "INVALID", "error": f"Token in invalid state: {token['status']}"}), 200
                cur.execute("""UPDATE meal_tokens SET status='redeemed', redeemed_by=%s, redeemed_at=NOW()
                               WHERE token_uid=%s AND status NOT IN ('redeemed','expired')""",
                            (scanner_id, token_uid))
                cur.execute("SELECT * FROM student_meals WHERE student_id = %s", (student_id,))
                student = cur.fetchone()
                cur.execute("SELECT * FROM meal_tokens WHERE token_uid = %s", (token_uid,))
                token = cur.fetchone()
        finally:
            conn.close()
        _log_scan(scanner_id, 'canteen_staff', 'token_qr', scanned_payload, student_id, token_uid, 'success')
        return jsonify({
            "status": "APPROVED", "token_uid": token_uid,
            "student": student,
            "meal_type": token['meal_type'],
            "redeemed_at": str(token['redeemed_at'])
        })
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/tokens/<token_uid>/qr-image', methods=['GET'])
@admin_bp.route('/api/tokens/<token_uid>/qr-image', methods=['GET'])
@authenticate
def get_token_qr_image(token_uid):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT token_qr_data FROM meal_tokens WHERE token_uid = %s", (token_uid,))
            token = cur.fetchone()
    finally:
        conn.close()
    if not token or not token.get('token_qr_data'):
        return jsonify({"error": "Token not found or QR not yet generated"}), 404
    buf = _qr_image(token['token_qr_data'])
    return send_file(buf, mimetype='image/png')

@admin_bp.route('/tokens', methods=['GET'])
@admin_bp.route('/api/tokens', methods=['GET'])
@authenticate
def list_tokens():
    conn = get_db()
    try:
        _lazy_expire_tokens(conn)
        with conn.cursor() as cur:
            query = "SELECT t.*, COALESCE(s.name, t.cached_student_name, t.student_id) as student_name, COALESCE(s.grade_section, 'Hostel') as grade_section FROM meal_tokens t LEFT JOIN student_meals s ON t.student_id = s.student_id WHERE 1=1"
            params = []
            user = _get_current_user()
            role = user.get('role')
            if role == 'student':
                sid = user.get('student_id')
                if not sid:
                    return jsonify([])
                query += " AND t.student_id = %s"
                params.append(sid)
            student_id = request.args.get('student_id')
            if student_id:
                query += " AND t.student_id = %s"
                params.append(student_id)
            meal_type = request.args.get('meal_type')
            if meal_type:
                query += " AND t.meal_type = %s"
                params.append(meal_type)
            status = request.args.get('status')
            if status:
                query += " AND t.status = %s"
                params.append(status)
            date_from = request.args.get('date_from')
            date_to = request.args.get('date_to')
            if date_from and date_to:
                query += " AND DATE(t.created_at) BETWEEN %s AND %s"
                params.extend([date_from, date_to])
            query += " ORDER BY t.created_at DESC"
            limit = request.args.get('limit', '50')
            query += f" LIMIT {int(limit)}"
            cur.execute(query, params)
            return jsonify(cur.fetchall())
    finally:
        conn.close()

@admin_bp.route('/tokens/recent', methods=['GET'])
@admin_bp.route('/api/tokens/recent', methods=['GET'])
@authenticate
@require_role('admin')
def recent_tokens():
    since = request.args.get('since')
    conn = get_db()
    try:
        with conn.cursor() as cur:
            if since:
                cur.execute("""SELECT t.*, s.name as student_name, s.grade_section
                               FROM meal_tokens t JOIN student_meals s ON t.student_id = s.student_id
                               WHERE t.created_at > %s OR t.redeemed_at > %s OR t.approved_at > %s
                               ORDER BY GREATEST(t.created_at, t.redeemed_at, t.approved_at) DESC LIMIT 50""",
                            (since, since, since))
            else:
                cur.execute("""SELECT t.*, s.name as student_name, s.grade_section
                               FROM meal_tokens t JOIN student_meals s ON t.student_id = s.student_id
                               ORDER BY t.created_at DESC LIMIT 50""")
            return jsonify(cur.fetchall())
    finally:
        conn.close()

@admin_bp.route('/tokens/student/active', methods=['GET'])
@admin_bp.route('/api/tokens/student/active', methods=['GET'])
@authenticate
@require_role('student')
def get_student_active_tokens():
    student_id = _get_current_user().get('student_id')
    if not student_id:
        return jsonify({"error": "No student profile linked"}), 400
    conn = get_db()
    try:
        _lazy_expire_tokens(conn)
        with conn.cursor() as cur:
            cur.execute("""SELECT * FROM meal_tokens
                           WHERE student_id = %s AND DATE(created_at) = CURDATE()
                           ORDER BY created_at DESC""", (student_id,))
            tokens = cur.fetchall()
            cur.execute("SELECT * FROM student_meals WHERE student_id = %s", (student_id,))
            student = cur.fetchone()
        return jsonify({
            "tokens": tokens, 
            "student": student, 
            "server_current_time": datetime.datetime.now().isoformat()
        })
    finally:
        conn.close()

# --- SCAN AUDIT LOG ---

@admin_bp.route('/scan-audit', methods=['GET'])
@admin_bp.route('/api/scan-audit', methods=['GET'])
@authenticate
@require_role('admin')
def get_scan_audit():
    limit = int(request.args.get('limit', '50'))
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM scan_audit_log ORDER BY created_at DESC LIMIT %s", (limit,))
            return jsonify(cur.fetchall())
    finally:
        conn.close()

# --- MEAL DISTRIBUTION LOG (existing style, mapped from token lifecycle) ---

@admin_bp.route('/meal-distribution', methods=['GET'])
@admin_bp.route('/api/meal-distribution', methods=['GET'])
@authenticate
@require_role('admin')
def get_meal_distribution():
    date_from = request.args.get('date_from') or request.args.get('date') or datetime.date.today().isoformat()
    date_to = request.args.get('date_to') or date_from
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("""SELECT t.*, s.name as student_name, s.grade_section
                           FROM meal_tokens t JOIN student_meals s ON t.student_id = s.student_id
                           WHERE DATE(t.created_at) BETWEEN %s AND %s AND t.status IN ('redeemed','token_issued','approved')
                           ORDER BY t.redeemed_at DESC, t.created_at DESC""", (date_from, date_to))
            return jsonify(cur.fetchall())
    finally:
        conn.close()

# --- EXISTING REGISTRATION MANAGEMENT (kept for backward compat) ---

def _strip_images(row):
    return {k: v for k, v in row.items() if k not in ('student_photo_base64', 'applicant_signature_base64', 'income_proof_base64')}

def _get_pending_registrations_from_local_file():
    candidates = [
        os.path.join(os.path.dirname(__file__), '..', 'registration_backend', 'uploads', 'pending_registrations.json'),
        os.path.join(os.path.dirname(__file__), 'uploads', 'pending_registrations.json'),
    ]
    for p in candidates:
        if os.path.exists(p):
            try:
                with open(p, 'r', encoding='utf-8') as f:
                    rows = json.load(f)
                    if isinstance(rows, list):
                        return rows
            except Exception:
                pass
    return []

@admin_bp.route('/registrations', methods=['GET'])
@admin_bp.route('/api/registrations', methods=['GET'])
@authenticate
@require_role('admin')
def get_all_registrations():
    try:
        conn = get_db()
        try:
            with conn.cursor() as cur:
                status_filter = request.args.get('status')
                if status_filter:
                    cur.execute("SELECT * FROM meal_registrations WHERE status = %s ORDER BY submitted_at DESC", (status_filter,))
                else:
                    cur.execute("SELECT * FROM meal_registrations ORDER BY submitted_at DESC")
                rows = cur.fetchall()
                for r in rows:
                    _normalize_registration_row(r)
        finally:
            conn.close()
        return jsonify({"registrations": rows, "count": len(rows)})
    except Exception as e:
        logger.warning(f"DB offline fallback for get_all_registrations: {e}")
        rows = _get_pending_registrations_from_local_file()
        status_filter = request.args.get('status')
        if status_filter:
            rows = [r for r in rows if r.get('status') == status_filter]
        return jsonify({"registrations": rows, "count": len(rows)})

@admin_bp.route('/database/students', methods=['GET'])
@admin_bp.route('/roster/students', methods=['GET'])
@authenticate
@require_role('admin')
def get_all_students_roster():
    try:
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM student_meals ORDER BY student_id ASC")
                rows = cur.fetchall()
                for r in rows:
                    r['degree_year'] = _format_degree_year(r.get('degree_year'))
        finally:
            conn.close()
        return jsonify({"students": rows, "count": len(rows)})
    except Exception as e:
        fallback_students = []
        return jsonify({"students": fallback_students, "count": len(fallback_students)})

@admin_bp.route('/students/promote-academic-year', methods=['POST'])
@admin_bp.route('/api/students/promote-academic-year', methods=['POST'])
@authenticate
@require_role('admin')
def promote_academic_year():
    """Batch migrates student academic years: 1st Year -> 2nd Year -> 3rd Year -> Graduated."""
    try:
        conn = get_db()
        try:
            _execute_year_migration(conn)
        finally:
            conn.close()
        _log_audit(_get_auditor_username(), 'MANUAL_YEAR_MIGRATION', 'student_meals', "Manual student academic year progression completed.")
        return jsonify({"message": "Student academic year progression completed successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _normalize_registration_row(r):
    if not r:
        return r
    if isinstance(r.get('submitted_at'), (datetime.datetime, datetime.date)):
        r['submitted_at'] = r['submitted_at'].isoformat()

    dept_no = str(r.get('dept_number') or r.get('dept_no') or r.get('last_year_id') or r.get('username') or r.get('app_no') or '').strip()
    student_name = quote_plus(str(r.get('student_name') or r.get('name') or 'Student').strip())
    avatar_url = f"https://ui-avatars.com/api/?name={student_name}&background=random"

    # Search for actual photo file in student_master_img
    resolved_photo = None
    possible_exts = ['.jpg', '.jpeg', '.png', '.JPG', '.JPEG', '.PNG']
    search_dirs = [
        (os.path.join(os.path.dirname(__file__), '..', 'registration_backend', 'uploads', 'student_master_img'), 'student_master_img'),
        (os.path.join(os.path.dirname(__file__), 'uploads', 'student_master_img'), 'student_master_img'),
    ]

    if dept_no:
        for ext in possible_exts:
            for dir_path, sub_folder in search_dirs:
                for candidate_name in [f"{dept_no}{ext}", f"{dept_no}_photo{ext}", f"{dept_no}_photo_safe{ext}"]:
                    full_path = os.path.join(dir_path, candidate_name)
                    if os.path.exists(full_path) and os.path.getsize(full_path) > 500:
                        resolved_photo = f"/uploads/{sub_folder}/{candidate_name}"
                        break
                if resolved_photo:
                    break
            if resolved_photo:
                break

    if not resolved_photo:
        existing = (r.get('student_image_path') or r.get('student_photo_url') or '').strip()
        if existing and (existing.startswith('http://') or existing.startswith('https://') or existing.startswith('data:')):
            resolved_photo = existing
        elif existing:
            rel_clean = existing.lstrip('/')
            check_bases = [
                os.path.join(os.path.dirname(__file__), '..', 'registration_backend'),
                os.path.join(os.path.dirname(__file__), '..'),
                os.path.dirname(__file__)
            ]
            file_valid = False
            for base in check_bases:
                fp = os.path.join(base, rel_clean)
                if os.path.exists(fp) and os.path.getsize(fp) > 500:
                    file_valid = True
                    break
            if file_valid:
                resolved_photo = '/' + rel_clean
            else:
                resolved_photo = avatar_url
        else:
            resolved_photo = avatar_url

    r['student_image_path'] = resolved_photo
    r['student_photo_url'] = resolved_photo

    sig = r.get('signature_path') or r.get('applicant_signature_url') or ''
    r['signature_path'] = sig
    r['applicant_signature_url'] = sig

    inc = r.get('income_proof_path') or r.get('income_proof_url') or ''
    r['income_proof_path'] = inc
    r['income_proof_url'] = inc

    return r

@admin_bp.route('/registrations/pending', methods=['GET'])
@admin_bp.route('/api/registrations/pending', methods=['GET'])
@authenticate
@require_role('admin')
def get_pending_registrations():
    try:
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM meal_registrations WHERE status = 'pending' ORDER BY submitted_at DESC")
                rows = cur.fetchall()
                for r in rows:
                    _normalize_registration_row(r)
        finally:
            conn.close()
        return jsonify({"registrations": rows, "count": len(rows)})
    except Exception as e:
        logger.warning(f"DB offline fallback for get_pending_registrations: {e}")
        rows = [r for r in _get_pending_registrations_from_local_file() if r.get('status', 'pending') == 'pending']
        return jsonify({"registrations": rows, "count": len(rows)})

@admin_bp.route('/registrations/<registration_id>', methods=['GET'])
@admin_bp.route('/api/registrations/<registration_id>', methods=['GET'])
@authenticate
@require_role('admin')
def get_registration(registration_id):
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM meal_registrations WHERE registration_id = %s", (registration_id,))
            target_row = cur.fetchone()
            if target_row:
                _normalize_registration_row(target_row)
    finally:
        conn.close()
    if not target_row:
        return jsonify({"error": "Registration not found"}), 404
    return jsonify({"registration": target_row})

@admin_bp.route('/registrations/<registration_id>/action', methods=['POST'])
@admin_bp.route('/api/registrations/<registration_id>/action', methods=['POST'])
@admin_bp.route('/registrations/<registration_id>/approve', methods=['POST'])
@admin_bp.route('/api/registrations/<registration_id>/approve', methods=['POST'])
@authenticate
@require_role('admin')
def act_on_registration(registration_id):
    try:
        data = request.json or {}
        action = data.get('action') or ('approve' if 'approve' in request.path else 'approve')
        if action not in ('approve', 'reject', 'forenoon', 'afternoon', 'both'):
            return jsonify({"error": "action must be 'approve', 'reject', 'forenoon', 'afternoon', or 'both'"}), 400
        
        conn = get_db()
        try:
            with conn.cursor() as cur:
                _ensure_student_meals_columns(cur)
                cur.execute("SELECT * FROM meal_registrations WHERE registration_id = %s", (registration_id,))
                target_row = cur.fetchone()
                if not target_row:
                    return jsonify({"error": "Registration not found"}), 404
                
                is_approve = action in ('approve', 'forenoon', 'afternoon', 'both')
                new_status = 'approved' if is_approve else 'rejected'
                
                # Determine meal attributes based on approval type
                if action == 'forenoon' or data.get('meal_type') == 'forenoon':
                    fn_meal, an_meal = 1, 0
                elif action == 'afternoon' or data.get('meal_type') == 'afternoon':
                    fn_meal, an_meal = 0, 1
                elif is_approve:
                    fn_meal, an_meal = 1, 1
                else:
                    fn_meal, an_meal = 0, 0

                cur.execute(
                    "UPDATE meal_registrations SET status = %s, forenoon_meal = %s, afternoon_meal = %s WHERE registration_id = %s",
                    (new_status, fn_meal, an_meal, registration_id)
                )
                target_row['status'] = new_status
                target_row['forenoon_meal'] = fn_meal
                target_row['afternoon_meal'] = an_meal

                email_sent = False
                if is_approve:
                    try:
                        sid = (target_row.get('dept_number') or target_row.get('app_no') or target_row.get('registration_id') or '').strip()
                        username = sid
                        raw_password = 'pass123'
                        pw_hash = bcrypt.hashpw(raw_password.encode('utf-8'), bcrypt.gensalt()).decode()
                        display_name = target_row.get('student_name') or username
                        student_email = target_row.get('email') or f"{username.lower()}@student.rkmvc"
                        grade_sec = f"{target_row.get('course', '')} - {target_row.get('department', '')}".strip(' -') or 'B.Sc. Comp Sci'
                        qr_secret = _gen_qr_secret(sid)

                        img_url = target_row.get('student_image_path') or target_row.get('student_photo_url') or f"https://ui-avatars.com/api/?name={quote_plus(display_name)}&background=random"
                        img_path = target_row.get('student_image_path') or target_row.get('student_photo_url') or ''
                        deg_year = target_row.get('degree_year') or target_row.get('year_of_degree') or '1st Year'
                        mobile_num = target_row.get('mobile_no') or target_row.get('phone') or 'N/A'

                        cur.execute("""
                            INSERT INTO student_meals (
                                student_id, username, email, password_hash, name, grade_section,
                                degree_year, mobile_no,
                                forenoon_meal, afternoon_meal, qr_secret, image_url, image_path,
                                student_image_path
                            ) VALUES (
                                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                            ) ON DUPLICATE KEY UPDATE
                                username=VALUES(username),
                                email=VALUES(email),
                                name=VALUES(name),
                                grade_section=VALUES(grade_section),
                                degree_year=IF(student_meals.degree_year IS NULL OR student_meals.degree_year = '' OR student_meals.degree_year = 'Enrolled', VALUES(degree_year), student_meals.degree_year),
                                mobile_no=VALUES(mobile_no),
                                forenoon_meal=VALUES(forenoon_meal),
                                afternoon_meal=VALUES(afternoon_meal),
                                qr_secret=VALUES(qr_secret),
                                image_url=VALUES(image_url),
                                image_path=VALUES(image_path),
                                student_image_path=VALUES(student_image_path)
                        """, (
                            sid, username, student_email, pw_hash, display_name, grade_sec,
                            deg_year, mobile_num,
                            fn_meal, an_meal, qr_secret, img_url, img_path,
                            img_path
                        ))

                        logger.info("Approved student registration credentials created in student_meals: %s", username)
                        target_email = (target_row.get('email') or target_row.get('student_email') or target_row.get('email_id') or '').strip()
                        if target_email:
                            try:
                                email_sent = email_service.send_credentials_email(
                                    target_email, username, raw_password, display_name
                                )
                            except Exception as email_err:
                                logger.warning("Approval email notification warning: %s", email_err)
                                email_sent = False
                    except Exception as app_err:
                        print("APPROVAL CRASH LOG:", str(app_err), flush=True)
                        logger.exception("APPROVAL CRASH LOG: %s", app_err)
                        return jsonify({"error": f"Failed to migrate student to student_meals: {str(app_err)}"}), 500
                elif action == 'reject':
                    student_email = (target_row.get('email') or target_row.get('student_email') or target_row.get('email_id') or '').strip()
                    display_name = target_row.get('student_name') or 'Applicant'
                    reject_reason = data.get('reason') or data.get('reject_reason')
                    if student_email:
                        try:
                            email_sent = email_service.send_rejection_email(student_email, display_name, reject_reason)
                        except Exception as rej_email_err:
                            logger.warning("Rejection email notification warning: %s", rej_email_err)
                            email_sent = False
        finally:
            conn.close()


        username_aud = getattr(request, 'user', {}).get('username', 'admin') if hasattr(request, 'user') and isinstance(getattr(request, 'user'), dict) else 'admin'
        _log_audit(username_aud, f'REGISTRATION_{new_status.upper()}', 'meal_registrations',
                   f"Registration '{registration_id}' for {target_row.get('student_name')} marked {new_status}")
        return jsonify({"success": True, "registration": target_row, "email_sent": email_sent})
    except Exception as e:
        print("APPROVAL CRASH LOG:", str(e), flush=True)
        logger.exception("APPROVAL CRASH LOG: %s", e)
        return jsonify({"error": f"Internal Error: {str(e)}"}), 500

# --- EXISTING IMPORT/EXPORT ROUTES (shortened, core functionality) ---

def _sanitize_for_json(obj):
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_sanitize_for_json(v) for v in obj]
    elif isinstance(obj, datetime.timedelta):
        total_seconds = int(obj.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    elif isinstance(obj, (datetime.datetime, datetime.date, datetime.time)):
        return str(obj)
    elif isinstance(obj, bytes):
        return obj.decode('utf-8', errors='ignore')
    return obj

def _lazy_expire_tokens(conn):
    try:
        with conn.cursor() as cur:
            cur.execute("DESCRIBE meal_tokens")
            cols = [c['Field'] for c in cur.fetchall()]
            time_col = 'expiry_time' if 'expiry_time' in cols else ('expires_at' if 'expires_at' in cols else None)
            create_col = 'created_at' if 'created_at' in cols else ('issued_at' if 'issued_at' in cols else None)
            
            where_clauses = []
            if time_col:
                where_clauses.append(f"({time_col} IS NOT NULL AND {time_col} < NOW())")
            if create_col:
                where_clauses.append(f"({create_col} IS NOT NULL AND TIMESTAMPDIFF(SECOND, {create_col}, NOW()) > 1800)")
            
            if where_clauses:
                query = f"""
                    UPDATE meal_tokens 
                    SET status = 'expired' 
                    WHERE status IN ('active', 'awaiting_scan', 'approved', 'token_issued', 'staff_verified')
                      AND ({' OR '.join(where_clauses)})
                """
                cur.execute(query)
    except Exception as e:
        logger.warning("Token auto-expire check warning: %s", e)
def _get_physical_mysql_tables(conn):
    """Refreshes and syncs physical MySQL tables for app_state without locking or blocking DB writes."""
    try:
        _lazy_expire_tokens(conn)
        with conn.cursor() as cur:
            cur.execute("SHOW TABLES")
            tables_in_db = [list(r.values())[0] for r in cur.fetchall()]

            cur.execute("SELECT data FROM app_state WHERE id = 1")
            row = cur.fetchone()
            db_data = json.loads(row['data']) if (row and row.get('data')) else {"tables": {}}
            tables_dict = db_data.setdefault('tables', {})

            for tbl in tables_in_db:
                if tbl in ('app_state', 'schema_migrations'):
                    continue

                cur.execute(f"DESCRIBE `{tbl}`")
                col_defs = cur.fetchall()
                columns_meta = []
                pk_name = None

                for c in col_defs:
                    col_name = c['Field']
                    col_type_str = str(c['Type']).upper()
                    is_pk = (c['Key'] == 'PRI') or (col_name in ['id', 'student_id', 'registration_id', 'token_uid', 'log_id'])
                    if is_pk and not pk_name:
                        pk_name = col_name

                    c_type = "TEXT"
                    if 'INT' in col_type_str or 'TINYINT' in col_type_str:
                        c_type = "INTEGER" if 'TINYINT(1)' not in col_type_str else "BOOLEAN"
                    elif 'DATE' in col_type_str or 'TIME' in col_type_str:
                        c_type = "DATETIME"

                    meta_item = {"name": col_name, "type": c_type, "nullable": c['Null'] == 'YES'}
                    if is_pk:
                        meta_item["primaryKey"] = True
                    columns_meta.append(meta_item)

                if not any(c.get('primaryKey') for c in columns_meta) and columns_meta:
                    columns_meta[0]['primaryKey'] = True

                cur.execute(f"SELECT * FROM `{tbl}`")
                raw_rows = cur.fetchall()
                formatted_rows = _sanitize_for_json(raw_rows)

                tables_dict[tbl] = {
                    "name": tbl,
                    "columns": columns_meta,
                    "rows": formatted_rows,
                    "recordCount": len(formatted_rows)
                }

            return db_data
    except Exception as e:
        logger.warning("Notice reading physical MySQL tables: %s", e)
        return None

@admin_bp.route('/tables', methods=['GET'])
@admin_bp.route('/api/tables', methods=['GET'])
@authenticate
def get_tables():
    conn = get_db()
    try:
        db_data = _get_physical_mysql_tables(conn)
        if not db_data:
            with conn.cursor() as cur:
                cur.execute("SELECT data FROM app_state WHERE id = 1")
                row = cur.fetchone()
                db_data = json.loads(row['data']) if row else {"tables": {}}
    finally:
        conn.close()

    tables_list = []
    for name, info in db_data.get('tables', {}).items():
        if name in ('app_state', 'schema_migrations'):
            continue
        tables_list.append({
            "name": name, "columns": info.get("columns", []),
            "recordCount": len(info.get("rows", [])),
            "createdAt": info.get("createdAt", "")
        })
    return jsonify(_sanitize_for_json(tables_list))

@admin_bp.route('/tables/<tableName>', methods=['GET'])
@admin_bp.route('/api/tables/<tableName>', methods=['GET'])
@authenticate
@limiter.limit("2000 per hour")
def get_table_records(tableName):
    # Direct physical MySQL query (fast, non-blocking, multi-threaded)
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SHOW TABLES")
            db_tables = [list(r.values())[0] for r in cur.fetchall()]
            matched_tbl = next((t for t in db_tables if t.lower() == tableName.lower()), None)
            
            if matched_tbl:
                cur.execute(f"DESCRIBE `{matched_tbl}`")
                col_defs = cur.fetchall()
                columns_meta = []
                for c in col_defs:
                    col_name = c['Field']
                    col_type_str = str(c['Type']).upper()
                    is_pk = (c['Key'] == 'PRI') or (col_name in ['id', 'student_id', 'registration_id', 'token_uid', 'log_id'])
                    c_type = "INTEGER" if ('INT' in col_type_str or 'TINYINT' in col_type_str) else ("DATETIME" if ('DATE' in col_type_str or 'TIME' in col_type_str) else "TEXT")
                    meta_item = {"name": col_name, "type": c_type, "nullable": c['Null'] == 'YES'}
                    if is_pk:
                        meta_item["primaryKey"] = True
                    columns_meta.append(meta_item)

                cur.execute(f"SELECT * FROM `{matched_tbl}`")
                raw_rows = cur.fetchall()
                formatted_rows = _sanitize_for_json(raw_rows)

                rows = list(formatted_rows)
                search = (request.args.get('search', '') or '').lower()
                if search:
                    rows = [r for r in rows if any(v is not None and search in str(v).lower() for v in r.values())]

                sort_by = request.args.get('sortBy', '')
                sort_order = request.args.get('sortOrder', 'asc')
                col_names = [c['name'] for c in columns_meta]
                if sort_by and sort_by in col_names:
                    def _sort_key(row):
                        v = row.get(sort_by)
                        if v is None:
                            return (1, '')
                        if isinstance(v, (int, float)):
                            return (0, v)
                        return (0, str(v).lower())
                    rows.sort(key=_sort_key, reverse=(sort_order == 'desc'))

                page = int(request.args.get('page', 1))
                limit = int(request.args.get('limit', 150))
                total = len(rows)
                total_pages = max(1, math.ceil(total / limit))
                paginated = rows[(page - 1) * limit: page * limit]

                return jsonify(_sanitize_for_json({
                    "tableName": matched_tbl,
                    "columns": columns_meta,
                    "rows": paginated,
                    "total": total,
                    "page": page,
                    "limit": limit,
                    "totalPages": total_pages,
                    "pagination": {
                        "page": page,
                        "limit": limit,
                        "totalPages": total_pages,
                        "totalRecords": total
                    }
                }))
    except Exception as err:
        logger.warning("Direct physical MySQL query fallback for %s: %s", tableName, err)
    finally:
        conn.close()

    # Secondary fallback to virtual tables in app_state if physical lookup fails
    conn = get_db()
    try:
        db_data = _get_physical_mysql_tables(conn)
        if not db_data:
            with conn.cursor() as cur:
                cur.execute("SELECT data FROM app_state WHERE id = 1")
                row = cur.fetchone()
                db_data = json.loads(row['data']) if row else {"tables": {}}
    finally:
        conn.close()

    tables_dict = db_data.get('tables', {})
    table = None
    for k, v in tables_dict.items():
        if k.lower() == tableName.lower():
            table = v
            break

    if not table:
        return jsonify({"error": f"Table '{tableName}' not found"}), 404

    rows = list(table.get('rows', []))
    search = (request.args.get('search', '') or '').lower()
    if search:
        rows = [r for r in rows if any(
            v is not None and search in str(v).lower()
            for v in r.values()
        )]

    sort_by = request.args.get('sortBy', '')
    sort_order = request.args.get('sortOrder', 'asc')
    col_names = [c['name'] for c in table.get('columns', [])]
    if sort_by and sort_by in col_names:
        def _sort_key(row):
            v = row.get(sort_by)
            if v is None:
                return (1, '')
            if isinstance(v, (int, float)):
                return (0, v)
            return (0, str(v).lower())
        rows.sort(key=_sort_key, reverse=(sort_order == 'desc'))

    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 150))
    total = len(rows)
    total_pages = max(1, math.ceil(total / limit))
    paginated = rows[(page - 1) * limit: page * limit]
    return jsonify(_sanitize_for_json({
        "tableName": tableName, "columns": table.get("columns", []),
        "rows": paginated,
        "pagination": {"page": page, "limit": limit, "totalPages": total_pages, "totalRecords": total}
    }))

def _coerce_value(val, col_type='TEXT'):
    if val is None:
        return None
    t = str(col_type).upper()
    if 'INT' in t or 'NUMBER' in t or 'BOOLEAN' in t or 'TINYINT' in t:
        if isinstance(val, bool):
            return 1 if val else 0
        if str(val).lower() in ['true', '1']:
            return 1
        if str(val).lower() in ['false', '0']:
            return 0
        try:
            return int(val)
        except (ValueError, TypeError):
            return val
    return str(val)

# --- CUSTOM TABLE CRUD (translated from server.ts) ---

def _get_auditor_username():
    from flask import g
    u = getattr(g, 'user', None)
    if isinstance(u, dict):
        return u.get('username') or u.get('student_id') or 'admin'
    if isinstance(u, str) and u:
        return u
    return 'admin'

@admin_bp.route('/tables', methods=['POST'])
@admin_bp.route('/api/tables', methods=['POST'])
@authenticate
def create_table():
    """Create a new custom table. Mirrors server.ts POST /api/tables (lines 228-269)."""
    try:
        body = request.json or {}
        name = body.get('name')
        columns = body.get('columns')
        if not name or not columns or not isinstance(columns, list):
            return jsonify({"error": "Table name and column schema are required"}), 400
        import re
        table_name_clean = re.sub(r'[^a-z0-9_]', '', name.strip().lower())
        if not table_name_clean:
            return jsonify({"error": "Invalid table name"}), 400
        # Validate and clean columns
        valid_columns = []
        for col in columns:
            col_name = re.sub(r'[^a-z0-9_]', '', str(col.get('name', '')).strip().lower())
            if not col_name:
                continue
            valid_columns.append({
                "name": col_name,
                "type": col.get('type', 'TEXT'),
                "primaryKey": bool(col.get('primaryKey', False)),
                "nullable": col.get('nullable', True) is not False
            })
        if not valid_columns:
            return jsonify({"error": "At least one valid column is required"}), 400
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT data FROM app_state WHERE id = 1 FOR UPDATE")
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "No app_state data"}), 500
                db_data = json.loads(row['data'])
                tables = db_data.setdefault('tables', {})
                if table_name_clean in tables:
                    return jsonify({"error": f"Table '{table_name_clean}' already exists"}), 400
                new_table = {
                    "createdAt": datetime.datetime.utcnow().isoformat() + "Z",
                    "columns": valid_columns,
                    "rows": []
                }
                tables[table_name_clean] = new_table
                cur.execute("UPDATE app_state SET data = %s WHERE id = 1", (json.dumps(db_data, default=str, ensure_ascii=False),))
            conn.commit()
        finally:
            conn.close()
        _log_audit(_get_auditor_username(), 'CREATE_TABLE', table_name_clean, f"Created table with {len(valid_columns)} columns")
        return jsonify({"message": f"Table '{table_name_clean}' created successfully", "table": new_table}), 201
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/drop-table', methods=['POST'])
@admin_bp.route('/api/drop-table', methods=['POST'])
@authenticate
def drop_table():
    """Drop a table. Mirrors server.ts POST /api/drop-table (lines 272-296)."""
    try:
        body = request.json or {}
        table_name = str(body.get('tableName', '')).strip()
        if not table_name:
            return jsonify({"error": "tableName is required"}), 400
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT data FROM app_state WHERE id = 1 FOR UPDATE")
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "No app_state data"}), 500
                db_data = json.loads(row['data'])
                tables = db_data.get('tables', {})
                # Case-insensitive lookup (matches server.ts line 279)
                table_key = next((k for k in tables if k.lower() == table_name.lower()), None)
                if not table_key:
                    return jsonify({"error": f"Table '{table_name}' not found"}), 404
                del tables[table_key]
                cur.execute("UPDATE app_state SET data = %s WHERE id = 1", (json.dumps(db_data, default=str, ensure_ascii=False),))
            conn.commit()
        finally:
            conn.close()
        _log_audit(_get_auditor_username(), 'DROP_TABLE', table_name, f"Dropped table '{table_name}'")
        return jsonify({"message": f"Table '{table_name}' dropped successfully"})
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/tables/<tableName>/records', methods=['POST'])
@admin_bp.route('/api/tables/<tableName>/records', methods=['POST'])
@authenticate
def insert_record(tableName):
    """Insert a record into a custom table. Mirrors server.ts POST /api/tables/:tableName/records (lines 298-341)."""
    try:
        body = request.json or {}
        conn = get_db()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT data FROM app_state WHERE id = 1 FOR UPDATE")
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "No app_state data"}), 500
                db_data = json.loads(row['data'])
                table = db_data.get('tables', {}).get(tableName)
                if not table:
                    return jsonify({"error": "Table not found"}), 404
                columns = table.get('columns', [])
                pk_column = next((c for c in columns if c.get('primaryKey')), None)
                new_record = {}
                for col in columns:
                    col_name = col['name']
                    input_val = body.get(col_name)
                    if input_val is None and col_name in ('served_at', 'timestamp', 'created_at'):
                        input_val = body.get('timestamp') or body.get('served_at') or body.get('created_at') or datetime.datetime.now().isoformat()
                    if col.get('primaryKey') and col.get('type') == 'NUMBER' and (input_val is None or input_val == ''):
                        # Auto increment numeric primary key
                        max_val = max((int(r.get(col_name) or 0) for r in table.get('rows', [])), default=0)
                        new_record[col_name] = max_val + 1
                    else:
                        if col.get('nullable') is False and (input_val is None or input_val == ''):
                            return jsonify({"error": f"Column '{col_name}' is not nullable but no value was provided"}), 400
                        new_record[col_name] = _coerce_value(input_val, col.get('type', 'TEXT'))

                for k, v in body.items():
                    if k not in new_record:
                        new_record[k] = _coerce_value(v, 'TEXT')
                # Check primary key uniqueness
                if pk_column:
                    pk_val = new_record.get(pk_column['name'])
                    if any(r.get(pk_column['name']) == pk_val for r in table.get('rows', [])):
                        return jsonify({"error": f"Duplicate primary key value '{pk_val}' for column '{pk_column['name']}'"}), 400
                table.setdefault('rows', []).append(new_record)

                if tableName.lower() == 'student_meals':
                    try:
                        from urllib.parse import quote_plus
                        sid = str(new_record.get('student_id', '')).strip()
                        sname = str(new_record.get('name', '')).strip()
                        sgrade = str(new_record.get('grade_section', 'B.Sc. Comp Sci')).strip()
                        fn_val = 1 if new_record.get('forenoon_meal') in [True, 1, 'true', '1', 'True'] else 0
                        an_val = 1 if new_record.get('afternoon_meal') in [True, 1, 'true', '1', 'True'] else 0
                        pw_hash = bcrypt.hashpw(b'pass123', bcrypt.gensalt()).decode()
                        qr_sec = _gen_qr_secret(sid)
                        img_url = f"https://ui-avatars.com/api/?name={quote_plus(sname)}&background=random"

                        cur.execute("""
                            INSERT INTO student_meals (student_id, username, email, password_hash, name, grade_section, forenoon_meal, afternoon_meal, qr_secret, image_url)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                name=VALUES(name),
                                grade_section=VALUES(grade_section),
                                forenoon_meal=VALUES(forenoon_meal),
                                afternoon_meal=VALUES(afternoon_meal)
                        """, (sid, sid, f"{sid.lower()}@student.rkmvc", pw_hash, sname, sgrade, fn_val, an_val, qr_sec, img_url))
                    except Exception as sql_err:
                        logger.warning("Error persisting student_meals to MySQL: %s", sql_err)

                cur.execute("UPDATE app_state SET data = %s WHERE id = 1", (json.dumps(db_data, default=str, ensure_ascii=False),))
            conn.commit()
        finally:
            conn.close()
        _log_audit(_get_auditor_username(), 'INSERT_RECORD', tableName, 'Inserted 1 record')
        return jsonify({"message": "Record added successfully", "record": new_record}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@admin_bp.route('/records/<record_id>', methods=['PUT'])
@admin_bp.route('/api/records/<record_id>', methods=['PUT'])
@authenticate
def update_record(record_id):
    """Update a record by primary key across physical SQL tables and app_state virtual tables."""
    try:
        body = request.get_json(silent=True) or {}
        target_table_name = body.get('targetTableName') or request.args.get('tableName')
        if not target_table_name:
            return jsonify({"error": "targetTableName is required in request body"}), 400
        conn = get_db()
        try:
            with conn.cursor() as cur:
                # 1. Physical SQL Table Update
                sql_sets = []
                params = []
                ignored_keys = {'targetTableName', 'tableName', 'created_at', 'updated_at'}

                for k, v in body.items():
                    if k in ignored_keys:
                        continue
                    if k in ('forenoon_meal', 'afternoon_meal'):
                        v = 1 if v in [True, 1, 'true', '1', 'True'] else 0
                    sql_sets.append(f"`{k}` = %s")
                    params.append(v)

                if sql_sets:
                    pk_field = 'student_id' if target_table_name.lower() == 'student_meals' else (
                        'registration_id' if target_table_name.lower() == 'meal_registrations' else (
                            'token_uid' if target_table_name.lower() == 'meal_tokens' else (
                                'id'
                            )
                        )
                    )
                    sql_query = f"UPDATE `{target_table_name}` SET {', '.join(sql_sets)} WHERE `{pk_field}` = %s"
                    params.append(str(record_id))
                    try:
                        cur.execute(sql_query, tuple(params))
                    except Exception as sql_err:
                        logger.warning("Error updating %s in MySQL: %s", target_table_name, sql_err)

                # 2. Synchronize Virtual app_state JSON store
                cur.execute("SELECT data FROM app_state WHERE id = 1 FOR UPDATE")
                row = cur.fetchone()
                if row and row.get('data'):
                    try:
                        db_data = json.loads(row['data'])
                        tables = db_data.get('tables', {})
                        if target_table_name in tables:
                            table = tables[target_table_name]
                            columns = table.get('columns', [])
                            pk_column = next((c for c in columns if c.get('primaryKey')), None)
                            pk_name = pk_column['name'] if pk_column else ('student_id' if target_table_name.lower() == 'student_meals' else ('registration_id' if target_table_name.lower() == 'meal_registrations' else (columns[0]['name'] if columns else 'id')))
                            rows = table.get('rows', [])
                            row_index = next((i for i, r in enumerate(rows) if str(r.get(pk_name, '')).strip().lower() == str(record_id).strip().lower()), -1)
                            
                            updated = dict(rows[row_index]) if row_index != -1 else {pk_name: str(record_id)}
                            for k, v in body.items():
                                if k not in ignored_keys:
                                    updated[k] = v
                                    
                            if row_index != -1:
                                rows[row_index] = updated
                            else:
                                rows.append(updated)
                                
                            cur.execute("UPDATE app_state SET data = %s WHERE id = 1", (json.dumps(db_data, default=str, ensure_ascii=False),))
                    except Exception as json_err:
                        logger.warning("app_state sync warning on update: %s", json_err)

            conn.commit()
        finally:
            conn.close()
        _log_audit(_get_auditor_username(), 'UPDATE_RECORD', target_table_name, f"Updated record ID: {record_id}")
        return jsonify({"message": "Record updated successfully", "record": body})
    except Exception as e:
        logger.error("Error in update_record: %s", e)
        return jsonify({"error": sani(e)}), 400

@admin_bp.route('/records/<record_id>', methods=['DELETE'])
@admin_bp.route('/api/records/<record_id>', methods=['DELETE'])
@admin_bp.route('/students/<record_id>', methods=['DELETE'])
@admin_bp.route('/api/students/<record_id>', methods=['DELETE'])
@admin_bp.route('/registrations/<record_id>', methods=['DELETE'])
@admin_bp.route('/api/registrations/<record_id>', methods=['DELETE'])
@admin_bp.route('/users/<record_id>', methods=['DELETE'])
@admin_bp.route('/api/users/<record_id>', methods=['DELETE'])
@authenticate
def delete_record(record_id):
    """Delete a record by primary key across physical MySQL tables and app_state."""
    try:
        req_json = request.get_json(silent=True) or {}
        target_table_name = request.args.get('tableName') or req_json.get('targetTableName') or req_json.get('tableName')
        if not target_table_name:
            if 'student' in request.path:
                target_table_name = 'student_meals'
            elif 'registration' in request.path:
                target_table_name = 'meal_registrations'
            elif 'user' in request.path:
                target_table_name = 'users'
            else:
                target_table_name = 'student_meals'
        
        conn = get_db()
        try:
            with conn.cursor() as cur:
                # Inspect schema to find real primary key column
                pk_field = None
                try:
                    cur.execute(f"DESCRIBE `{target_table_name}`")
                    cols_info = cur.fetchall()
                    for c in cols_info:
                        if c.get('Key') == 'PRI':
                            pk_field = c.get('Field')
                            break
                except Exception:
                    pass

                if not pk_field:
                    pk_field = 'student_id' if target_table_name.lower() == 'student_meals' else (
                        'registration_id' if target_table_name.lower() == 'meal_registrations' else (
                            'token_uid' if target_table_name.lower() == 'meal_tokens' else (
                                'log_id' if target_table_name.lower() == 'meal_distribution_log' else 'id'
                            )
                        )
                    )

                # Physical SQL Delete
                try:
                    cur.execute(f"DELETE FROM `{target_table_name}` WHERE `{pk_field}` = %s", (record_id,))
                except Exception as sql_del_err:
                    logger.warning("SQL primary key deletion warning for table %s: %s", target_table_name, sql_del_err)
                    try:
                        cur.execute(f"DELETE FROM `{target_table_name}` WHERE id = %s", (record_id,))
                    except Exception:
                        pass

                # Cascade deletes
                if target_table_name.lower() == 'student_meals':
                    try:
                        cur.execute("DELETE FROM meal_registrations WHERE dept_number = %s OR app_no = %s OR registration_id = %s", (record_id, record_id, record_id))
                    except Exception:
                        pass
                    try:
                        cur.execute("DELETE FROM meal_tokens WHERE student_id = %s", (record_id,))
                    except Exception:
                        pass
                elif target_table_name.lower() == 'meal_registrations':
                    try:
                        cur.execute("DELETE FROM student_meals WHERE student_id = %s OR registration_id = %s", (record_id, record_id))
                    except Exception:
                        pass
                    try:
                        cur.execute("DELETE FROM meal_tokens WHERE student_id = %s", (record_id,))
                    except Exception:
                        pass

            conn.commit()
            
            # Immediately resync physical MySQL tables with app_state JSON store
            _get_physical_mysql_tables(conn)
        finally:
            conn.close()

        _log_audit(_get_auditor_username(), 'DELETE_RECORD', target_table_name, f"Deleted record ID: {record_id}")
        return jsonify({"success": True, "message": "Record deleted successfully"})
    except Exception as e:
        print("DELETE CRASH LOG:", str(e), flush=True)
        logger.exception("Error in delete_record: %s", e)
        return jsonify({"error": f"Failed to delete record: {str(e)}"}), 500


@admin_bp.route('/import/preview', methods=['POST'])
@admin_bp.route('/api/import/preview', methods=['POST'])
@authenticate
def import_preview():
    try:
        uploaded_file = request.files.get('file')
        if not uploaded_file:
            return jsonify({"error": "No file uploaded"}), 400
        content = uploaded_file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
        columns = [{"name": col, "type": "TEXT"} for col in (reader.fieldnames or [])]
        preview_rows = rows[:5]
        for row in preview_rows:
            for k, v in row.items():
                if v == '':
                    row[k] = None
        return jsonify({"filename": uploaded_file.filename, "columns": columns, "preview": preview_rows, "totalRows": len(rows)})
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/import/csv', methods=['POST'])
@admin_bp.route('/api/import/csv', methods=['POST'])
@authenticate
def import_csv():
    try:
        uploaded_file = request.files.get('file')
        target_table = request.form.get('targetTable')
        columns_mapping = request.form.get('columnsMapping')
        duplicate_option = request.form.get('duplicateOption', 'skip')
        if not uploaded_file or not target_table:
            return jsonify({"error": "File and target table required"}), 400
        content = uploaded_file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        mapping = json.loads(columns_mapping) if columns_mapping else None
        raw_rows = list(reader)
        conn = get_db()
        inserted = 0
        updated = 0
        skipped = 0
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT data FROM app_state WHERE id = 1 FOR UPDATE")
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "No app_state data"}), 500
                db_data = json.loads(row['data'])
                tables = db_data.get('tables', {})
                if target_table not in tables:
                    columns = [{"name": col, "type": "TEXT"} for col in (reader.fieldnames or [])]
                    tables[target_table] = {"columns": columns, "rows": []}
                table = tables[target_table]
                col_names = [c['name'] for c in table.get('columns', [])]
                for raw_row in raw_rows:
                    cleaned = {k: v for k, v in raw_row.items() if k is not None}
                    for k, v in cleaned.items():
                        if v == '':
                            cleaned[k] = None
                    if mapping:
                        db_row = {}
                        for csv_key, db_col in mapping.items():
                            if db_col and csv_key in cleaned:
                                db_row[db_col] = cleaned[csv_key]
                    else:
                        db_row = cleaned
                    if not any(v is not None for v in db_row.values()):
                        skipped += 1
                        continue
                    pk_val = db_row.get(col_names[0]) if col_names else None
                    existing_idx = -1
                    if pk_val is not None:
                        for i, r in enumerate(table['rows']):
                            if r.get(col_names[0]) == pk_val:
                                existing_idx = i
                                break
                    if existing_idx >= 0:
                        if duplicate_option == 'update':
                            table['rows'][existing_idx].update(db_row)
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        table['rows'].append(db_row)
                        inserted += 1
                db_data['tables'] = tables
                cur.execute("UPDATE app_state SET data = %s WHERE id = 1", (json.dumps(db_data),))
            import_log_id = 'imp_' + secrets.token_hex(6)
            with conn.cursor() as cur:
                cur.execute("INSERT INTO import_logs (id, filename, records_imported, status, created_at) VALUES (%s, %s, %s, %s, NOW())",
                    (import_log_id, uploaded_file.filename, inserted + updated,
                     'SUCCESS' if inserted + updated > 0 else ('PARTIAL' if skipped > 0 and inserted + updated > 0 else 'FAILED')))
            conn.commit()
        finally:
            conn.close()
        _log_audit(_get_auditor_username(), 'IMPORT_CSV', target_table, f"Imported {inserted} inserted, {updated} updated, {skipped} skipped")
        return jsonify({"message": "Import completed", "summary": {"totalCsvRows": len(raw_rows), "inserted": inserted, "updated": updated, "skipped": skipped}})
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/export/csv', methods=['POST'])
@admin_bp.route('/api/export/csv', methods=['POST'])
@authenticate
def export_csv():
    try:
        data = request.json or {}
        table_name = data.get('tableName')
        selected_columns = data.get('columns')
        conn = get_db()
        try:
            db = _get_physical_mysql_tables(conn)
            if db is None:
                with conn.cursor() as cur:
                    cur.execute("SELECT data FROM app_state WHERE id = 1")
                    row = cur.fetchone()
                db = json.loads(row['data']) if row else None
        finally:
            conn.close()
        if not db:
            return jsonify({"error": "No data"}), 404
        table = db.get('tables', {}).get(table_name)
        if not table:
            return jsonify({"error": "Table not found"}), 404
        rows = table.get('rows', [])
        cols = selected_columns if selected_columns else [c['name'] for c in table.get('columns', [])]
        csv_content = io.StringIO()
        if rows:
            w = csv.DictWriter(csv_content, fieldnames=cols, extrasaction='ignore')
            w.writeheader()
            filtered = [{k: r.get(k) for k in cols} for r in rows]
            w.writerows(filtered)
        export_log_id = 'exp_' + secrets.token_hex(6)
        conn2 = get_db()
        try:
            with conn2.cursor() as cur:
                cur.execute("INSERT INTO export_logs (id, filename, records_exported, format, created_at) VALUES (%s, %s, %s, 'csv', NOW())",
                    (export_log_id, f"{table_name}_export.csv", len(rows)))
            conn2.commit()
        finally:
            conn2.close()
        _log_audit(_get_auditor_username(), 'EXPORT_CSV', table_name, f"Exported {len(rows)} records to CSV")
        mem = io.BytesIO(csv_content.getvalue().encode('utf-8'))
        mem.seek(0)
        return send_file(mem, mimetype="text/csv", as_attachment=True, download_name=f"{table_name}_export.csv")
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/export/excel', methods=['POST'])
@admin_bp.route('/api/export/excel', methods=['POST'])
@authenticate
def export_excel():
    try:
        data = request.json or {}
        table_name = data.get('tableName')
        selected_columns = data.get('columns')
        conn = get_db()
        try:
            db = _get_physical_mysql_tables(conn)
            if db is None:
                with conn.cursor() as cur:
                    cur.execute("SELECT data FROM app_state WHERE id = 1")
                    row = cur.fetchone()
                db = json.loads(row['data']) if row else None
        finally:
            conn.close()
        if not db:
            return jsonify({"error": "No data"}), 404
        table = db.get('tables', {}).get(table_name)
        if not table:
            return jsonify({"error": "Table not found"}), 404
        rows = table.get('rows', [])
        cols = selected_columns if selected_columns else [c['name'] for c in table.get('columns', [])]
        filtered = [{k: r.get(k) for k in cols} for r in rows]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name[:31]
        if cols:
            ws.append(cols)
        for r in filtered:
            ws.append([r.get(c) for c in cols])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        export_log_id = 'exp_' + secrets.token_hex(6)
        conn2 = get_db()
        try:
            with conn2.cursor() as cur:
                cur.execute("INSERT INTO export_logs (id, filename, records_exported, format, created_at) VALUES (%s, %s, %s, 'excel', NOW())",
                    (export_log_id, f"{table_name}_export.xlsx", len(rows)))
            conn2.commit()
        finally:
            conn2.close()
        _log_audit(_get_auditor_username(), 'EXPORT_EXCEL', table_name, f"Exported {len(rows)} records to Excel")
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"{table_name}_export.xlsx")
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/export/json', methods=['POST'])
@admin_bp.route('/api/export/json', methods=['POST'])
@authenticate
def export_json():
    try:
        data = request.json or {}
        table_name = data.get('tableName')
        selected_columns = data.get('columns')
        conn = get_db()
        try:
            db = _get_physical_mysql_tables(conn)
            if db is None:
                with conn.cursor() as cur:
                    cur.execute("SELECT data FROM app_state WHERE id = 1")
                    row = cur.fetchone()
                db = json.loads(row['data']) if row else None
        finally:
            conn.close()
        if not db:
            return jsonify({"error": "No data"}), 404
        table = db.get('tables', {}).get(table_name)
        if not table:
            return jsonify({"error": "Table not found"}), 404
        rows = table.get('rows', [])
        cols = selected_columns if selected_columns else [c['name'] for c in table.get('columns', [])]
        filtered = [{k: r.get(k) for k in cols} for r in rows]
        export_log_id = 'exp_' + secrets.token_hex(6)
        conn2 = get_db()
        try:
            with conn2.cursor() as cur:
                cur.execute("INSERT INTO export_logs (id, filename, records_exported, format, created_at) VALUES (%s, %s, %s, 'json', NOW())",
                    (export_log_id, f"{table_name}_export.json", len(rows)))
            conn2.commit()
        finally:
            conn2.close()
        _log_audit(_get_auditor_username(), 'EXPORT_JSON', table_name, f"Exported {len(rows)} records to JSON")
        return send_file(io.BytesIO(json.dumps(filtered, indent=2).encode('utf-8')),
                         mimetype="application/json", as_attachment=True, download_name=f"{table_name}_export.json")
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/convert/csv-to-json', methods=['POST'])
@admin_bp.route('/api/convert/csv-to-json', methods=['POST'])
@authenticate
def convert_csv_to_json():
    try:
        uploaded_file = request.files.get('file')
        if not uploaded_file:
            return jsonify({"error": "No file uploaded"}), 400
        content = uploaded_file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        return jsonify(list(reader))
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/convert/csv-to-excel', methods=['POST'])
@admin_bp.route('/api/convert/csv-to-excel', methods=['POST'])
@authenticate
def convert_csv_to_excel():
    try:
        uploaded_file = request.files.get('file')
        if not uploaded_file:
            return jsonify({"error": "No file uploaded"}), 400
        content = uploaded_file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
        wb = openpyxl.Workbook()
        ws = wb.active
        if rows:
            ws.append(list(rows[0].keys()))
            for r in rows:
                ws.append(list(r.values()))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        name = uploaded_file.filename.rsplit('.', 1)[0] + '.xlsx' if uploaded_file.filename else 'output.xlsx'
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=name)
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/convert/json-to-csv', methods=['POST'])
@admin_bp.route('/api/convert/json-to-csv', methods=['POST'])
@authenticate
def convert_json_to_csv():
    try:
        uploaded_file = request.files.get('file')
        if not uploaded_file:
            return jsonify({"error": "No file uploaded"}), 400
        data = json.loads(uploaded_file.read().decode('utf-8'))
        if not isinstance(data, list) or not data:
            return jsonify({"error": "Empty or invalid JSON array"}), 400
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=data[0].keys())
        w.writeheader()
        w.writerows(data)
        name = uploaded_file.filename.rsplit('.', 1)[0] + '.csv' if uploaded_file.filename else 'output.csv'
        return send_file(io.BytesIO(buf.getvalue().encode('utf-8')), mimetype="text/csv",
                         as_attachment=True, download_name=name)
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/convert/excel-to-csv', methods=['POST'])
@admin_bp.route('/api/convert/excel-to-csv', methods=['POST'])
@authenticate
def convert_excel_to_csv():
    try:
        uploaded_file = request.files.get('file')
        if not uploaded_file:
            return jsonify({"error": "No file uploaded"}), 400
        buf = io.BytesIO(uploaded_file.read())
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "Empty spreadsheet"}), 400
        csv_buf = io.StringIO()
        w = csv.writer(csv_buf)
        for row in rows:
            w.writerow(row)
        name = uploaded_file.filename.rsplit('.', 1)[0] + '.csv' if uploaded_file.filename else 'output.csv'
        return send_file(io.BytesIO(csv_buf.getvalue().encode('utf-8')), mimetype="text/csv",
                         as_attachment=True, download_name=name)
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/convert/csv-to-sql', methods=['POST'])
@admin_bp.route('/api/convert/csv-to-sql', methods=['POST'])
@authenticate
def convert_csv_to_sql():
    try:
        uploaded_file = request.files.get('file')
        table_name = request.form.get('tableName', 'imported_table')
        if not uploaded_file:
            return jsonify({"error": "No file uploaded"}), 400
        content = uploaded_file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
        if not rows:
            return jsonify({"error": "Empty CSV"}), 400
        cols = list(rows[0].keys())
        lines = []
        for r in rows:
            vals = []
            for c in cols:
                v = r.get(c, '')
                if v is None or v == '':
                    vals.append('NULL')
                else:
                    escaped = str(v).replace("'", "''")
                    vals.append(f"'{escaped}'")
            lines.append(f"INSERT INTO {table_name} ({', '.join(cols)}) VALUES ({', '.join(vals)});")
        sql = '\n'.join(lines)
        name = f"{table_name}_inserts.sql"
        return send_file(io.BytesIO(sql.encode('utf-8')), mimetype="text/plain",
                         as_attachment=True, download_name=name)
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

@admin_bp.route('/logs/audit', methods=['GET'])
@admin_bp.route('/api/logs/audit', methods=['GET'])
@admin_bp.route('/system/audit-logs', methods=['GET'])
@admin_bp.route('/api/system/audit-logs', methods=['GET'])
@authenticate
def get_audit_logs():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM audit_logs ORDER BY created_at DESC LIMIT 100")
            rows = cur.fetchall()
            for r in rows:
                if isinstance(r.get('created_at'), (datetime.datetime, datetime.date)):
                    r['created_at'] = r['created_at'].isoformat()
            return jsonify(rows)
    finally:
        conn.close()

@admin_bp.route('/logs/imports', methods=['GET'])
@admin_bp.route('/api/logs/imports', methods=['GET'])
@authenticate
def get_import_logs():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM import_logs ORDER BY created_at DESC LIMIT 50")
            return jsonify(cur.fetchall())
    finally:
        conn.close()

@admin_bp.route('/logs/exports', methods=['GET'])
@admin_bp.route('/api/logs/exports', methods=['GET'])
@authenticate
def get_export_logs():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM export_logs ORDER BY created_at DESC LIMIT 50")
            return jsonify(cur.fetchall())
    finally:
        conn.close()

# --- REGISTRATION PROXY ---
# Forwards the public meal registration form to the registration Flask backend
@admin_bp.route('/register', methods=['POST'])
@admin_bp.route('/api/register', methods=['POST'])
def proxy_register():
    try:
        reg_url = f"{REGISTRATION_BACKEND_URL}/api/register"
        _files = {}
        for key in request.files:
            f = request.files[key]
            _files[key] = (f.filename, f.stream, f.content_type or 'application/octet-stream')
        resp = requests.post(reg_url, data=dict(request.form), files=_files, timeout=300)
        return (resp.content, resp.status_code, dict(resp.headers))
    except requests.ConnectionError:
        return jsonify({"error": "Registration service is offline. Contact administrator."}), 503

# --- COMMUNICATIONS (Broadcast Email) ---

@admin_bp.route('/communications/students', methods=['GET'])
@admin_bp.route('/api/communications/students', methods=['GET'])
@authenticate
@require_role('admin', 'approval_staff', 'canteen_staff')
def comm_students_with_email():
    conn = get_db()
    try:
        grade_section = request.args.get('grade_section')
        forenoon = request.args.get('forenoon')
        afternoon = request.args.get('afternoon')
        q = request.args.get('q')
        query = """
            SELECT s.student_id, s.name, s.grade_section, s.forenoon_meal, s.afternoon_meal,
                   COALESCE(NULLIF(s.email, ''), CONCAT(LOWER(s.student_id), '@student.rkmvc')) as email,
                   s.name as display_name
            FROM student_meals s
            WHERE 1=1
        """
        params = []
        if grade_section:
            query += " AND s.grade_section = %s"
            params.append(grade_section)
        if forenoon in ('1', 'true', 'True'):
            query += " AND s.forenoon_meal = 1"
        if afternoon in ('1', 'true', 'True'):
            query += " AND s.afternoon_meal = 1"
        if q:
            query += " AND (s.name LIKE %s OR s.student_id LIKE %s OR s.email LIKE %s)"
            params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
        query += " ORDER BY s.student_id ASC"
        with conn.cursor() as cur:
            cur.execute(query, params)
            students = cur.fetchall()
            for st in students:
                st['forenoon_meal'] = 1 if st.get('forenoon_meal') in (True, 1, '1', 'true') else 0
                st['afternoon_meal'] = 1 if st.get('afternoon_meal') in (True, 1, 'true', '1', 'true') else 0
            return jsonify(students)
    finally:
        conn.close()

@admin_bp.route('/communications/filter-options', methods=['GET'])
@admin_bp.route('/api/communications/filter-options', methods=['GET'])
@authenticate
@require_role('admin', 'approval_staff', 'canteen_staff')
def comm_filter_options():
    conn = get_db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT grade_section FROM student_meals WHERE grade_section IS NOT NULL AND grade_section != '' ORDER BY grade_section")
            grade_sections = [r['grade_section'] for r in cur.fetchall()]
            cur.execute("SELECT COUNT(*) as c FROM student_meals")
            total = cur.fetchone()['c']
            cur.execute("SELECT COUNT(*) as c FROM student_meals WHERE email IS NOT NULL AND email != ''")
            with_email = cur.fetchone()['c']
            if with_email == 0 and total > 0:
                with_email = total
        return jsonify({"grade_sections": grade_sections, "total_students": total, "with_email": with_email})
    finally:
        conn.close()

@admin_bp.route('/communications/send', methods=['POST'])
@admin_bp.route('/api/communications/send', methods=['POST'])
@authenticate
@require_role('admin', 'approval_staff', 'canteen_staff')
def comm_send_broadcast():
    try:
        data = request.json or {}
        student_ids = data.get('student_ids', [])
        subject = data.get('subject', '').strip()
        body = data.get('body', '').strip()
        sender_email = data.get('sender_email', '').strip() or None
        if not student_ids or not subject or not body:
            return jsonify({"error": "student_ids, subject, and body are required"}), 400
        conn = get_db()
        try:
            with conn.cursor() as cur:
                placeholders = ','.join(['%s'] * len(student_ids))
                cur.execute(f"""
                    SELECT student_id, name, COALESCE(NULLIF(email, ''), CONCAT(LOWER(student_id), '@student.rkmvc')) as email
                    FROM student_meals
                    WHERE student_id IN ({placeholders})
                """, student_ids)
                recipients = [{"email": r['email'], "name": r['name']} for r in cur.fetchall()]
        finally:
            conn.close()
        if not recipients:
            return jsonify({"error": "No recipients found for selected student IDs"}), 400
        result = email_service.send_broadcast_email(recipients, subject, body, sender_email)
        _log_audit(_get_auditor_username(), 'BROADCAST_EMAIL', 'communications',
                   json.dumps({"subject": subject, "recipient_count": len(recipients),
                                "sent": result['sent'], "failed": result['failed'],
                                "sender_email": sender_email or email_service.SMTP_FROM}))
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": sani(e)}), 500

# --- MEAL REPORT GENERATION ---

@admin_bp.route('/meal-report/generate', methods=['POST'])
@admin_bp.route('/api/meal-report/generate', methods=['POST'])
@authenticate
@require_role('admin')
def generate_meal_report():
    """Generate a PDF meal report for a given date and meal type (forenoon/afternoon).
    Called automatically by the frontend at the end of each meal window."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        import io

        data = request.json or {}
        meal_type = data.get('meal_type', '').lower()   # 'forenoon' or 'afternoon'
        report_date = data.get('date') or datetime.date.today().isoformat()

        if meal_type not in ('forenoon', 'afternoon'):
            return jsonify({"error": "meal_type must be 'forenoon' or 'afternoon'"}), 400

        conn = get_db()
        try:
            with conn.cursor() as cur:
                # Fetch all tokens for this meal type and date
                cur.execute("""
                    SELECT
                        t.token_uid, t.meal_type, t.status,
                        t.created_at, t.redeemed_at, t.redeemed_by,
                        s.name AS student_name, s.grade_section,
                        COALESCE(s.register_number, s.student_id) AS reg_no
                    FROM meal_tokens t
                    JOIN student_meals s ON t.student_id = s.student_id
                    WHERE t.meal_type = %s AND DATE(t.created_at) = %s
                    ORDER BY t.status DESC, t.created_at ASC
                """, (meal_type, report_date))
                rows = cur.fetchall()

                # Summary counts
                total = len(rows)
                redeemed = sum(1 for r in rows if r['status'] == 'redeemed')
                not_redeemed = sum(1 for r in rows if r['status'] != 'redeemed')
                expired = sum(1 for r in rows if r['status'] == 'expired')
                rejected = sum(1 for r in rows if r['status'] == 'rejected')
        finally:
            conn.close()

        # Build PDF in memory
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=18*mm, bottomMargin=18*mm)

        styles = getSampleStyleSheet()
        SAFFRON = colors.Color(1.0, 0.6, 0.2)
        SLATE900 = colors.Color(0.059, 0.090, 0.165)
        SLATE500 = colors.Color(0.392, 0.455, 0.545)
        SLATE100 = colors.Color(0.945, 0.957, 0.976)

        title_style = ParagraphStyle('Title', fontSize=15, fontName='Helvetica-Bold',
                                     textColor=SLATE900, spaceAfter=2)
        sub_style = ParagraphStyle('Sub', fontSize=8.5, fontName='Helvetica',
                                   textColor=SLATE500, spaceAfter=4)
        label_style = ParagraphStyle('Label', fontSize=7, fontName='Helvetica-Bold',
                                     textColor=SLATE500, spaceAfter=1)
        stat_style = ParagraphStyle('Stat', fontSize=18, fontName='Helvetica-Bold',
                                    textColor=SAFFRON)

        meal_label = 'Morning (Forenoon)' if meal_type == 'forenoon' else 'Afternoon'
        generated_at = datetime.datetime.now().strftime('%d %b %Y %I:%M %p')
        report_date_fmt = datetime.datetime.strptime(report_date, '%Y-%m-%d').strftime('%d %B %Y')

        story = []

        # Header bar (via a 1-row table for coloured stripe effect)
        header_data = [[Paragraph(
            f'<b>RKMVC MEALFLOW</b> &nbsp;|&nbsp; {meal_label} Meal Report &nbsp;|&nbsp; {report_date_fmt}',
            ParagraphStyle('H', fontSize=9, fontName='Helvetica-Bold', textColor=colors.white)
        )]]
        header_tbl = Table(header_data, colWidths=[180*mm])
        header_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), SLATE900),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(header_tbl)
        story.append(Spacer(1, 6*mm))

        story.append(Paragraph(f'{meal_label} Meal Distribution Report', title_style))
        story.append(Paragraph(f'Generated automatically at {generated_at} &nbsp;·&nbsp; Date: {report_date_fmt}', sub_style))
        story.append(HRFlowable(width='100%', thickness=1.5, color=SAFFRON, spaceAfter=5*mm))

        # Summary stats row
        stat_cells = [
            [Paragraph('TOTAL TOKENS', label_style), Paragraph('REDEEMED', label_style),
             Paragraph('NOT REDEEMED', label_style), Paragraph('EXPIRED', label_style), Paragraph('REJECTED', label_style)],
            [Paragraph(str(total), stat_style), Paragraph(str(redeemed), ParagraphStyle('S2', fontSize=18, fontName='Helvetica-Bold', textColor=colors.Color(0.133,0.545,0.133))),
             Paragraph(str(not_redeemed), ParagraphStyle('S3', fontSize=18, fontName='Helvetica-Bold', textColor=colors.Color(0.8,0.4,0.1))),
             Paragraph(str(expired), ParagraphStyle('S4', fontSize=18, fontName='Helvetica-Bold', textColor=SLATE500)),
             Paragraph(str(rejected), ParagraphStyle('S5', fontSize=18, fontName='Helvetica-Bold', textColor=colors.Color(0.8,0.1,0.1)))]
        ]
        col_w = 180*mm / 5
        stat_tbl = Table(stat_cells, colWidths=[col_w]*5)
        stat_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), SLATE100),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('ROUNDEDCORNERS', [4]),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.Color(0.878, 0.902, 0.933)),
        ]))
        story.append(stat_tbl)
        story.append(Spacer(1, 6*mm))

        # Detail table
        col_headers = ['#', 'Reg. No.', 'Student Name', 'Section', 'Status', 'Redeemed At', 'Served By']
        col_widths = [8*mm, 30*mm, 45*mm, 32*mm, 22*mm, 28*mm, 22*mm]
        table_data = [col_headers]

        STATUS_COLORS = {
            'redeemed': colors.Color(0.133, 0.545, 0.133),
            'expired': SLATE500,
            'rejected': colors.Color(0.8, 0.1, 0.1),
            'token_issued': colors.Color(0.1, 0.4, 0.8),
            'approved': colors.Color(0.1, 0.6, 0.5),
        }

        style_map = {}
        for i, row in enumerate(rows):
            redeemed_at = ''
            if row.get('redeemed_at'):
                try:
                    redeemed_at = str(row['redeemed_at'])
                    if 'T' in redeemed_at or ' ' in redeemed_at:
                        dt = datetime.datetime.fromisoformat(str(row['redeemed_at']))
                        redeemed_at = dt.strftime('%I:%M %p')
                except Exception:
                    redeemed_at = str(row['redeemed_at'])
            table_data.append([
                str(i + 1),
                row.get('reg_no') or '',
                row.get('student_name') or '',
                row.get('grade_section') or '',
                (row.get('status') or '').replace('_', ' ').title(),
                redeemed_at,
                row.get('redeemed_by') or '',
            ])
            status_key = (row.get('status') or '').lower()
            style_map[i + 1] = STATUS_COLORS.get(status_key, SLATE500)

        detail_tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl_style = [
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), SLATE900),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7.5),
            ('TOPPADDING', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, SLATE100]),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.Color(0.878, 0.902, 0.933)),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ]
        # Color-code status column
        for row_idx, col in style_map.items():
            tbl_style.append(('TEXTCOLOR', (4, row_idx), (4, row_idx), col))
            tbl_style.append(('FONTNAME', (4, row_idx), (4, row_idx), 'Helvetica-Bold'))

        detail_tbl.setStyle(TableStyle(tbl_style))
        story.append(detail_tbl)

        story.append(Spacer(1, 8*mm))
        story.append(HRFlowable(width='100%', thickness=0.5, color=SLATE100))
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(
            f'End of {meal_label} session report &nbsp;·&nbsp; RKMVC MealFlow Dining System &nbsp;·&nbsp; {generated_at}',
            ParagraphStyle('Footer', fontSize=7, fontName='Helvetica', textColor=SLATE500, alignment=TA_CENTER)
        ))

        doc.build(story)
        buf.seek(0)

        filename = f"meal_report_{meal_type}_{report_date}.pdf"
        _log_audit(_get_auditor_username(), 'GENERATE_MEAL_REPORT', 'meal_tokens',
                   f"Auto-generated {meal_label} report for {report_date}: {redeemed}/{total} redeemed")

        return send_file(buf, mimetype='application/pdf',
                         as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({"error": sani(e)}), 500


# --- HEALTH CHECK (no auth) ---

@admin_bp.route('/health', methods=['GET'])
@admin_bp.route('/api/health', methods=['GET'])
def health():
    try:
        conn = get_db()
        conn.close()
        return jsonify({"status": "healthy", "db": "connected"})
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return jsonify({"status": "unhealthy", "db": "disconnected"}), 503

# Register blueprints
# from blueprints import register_blueprints
# register_blueprints(app)

@admin_bp.route('/uploads/<path:filename>')
def serve_uploads(filename):
    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'registration_backend', 'uploads')
    if not os.path.exists(uploads_dir):
        uploads_dir = 'uploads'
    return send_from_directory(uploads_dir, filename)

@admin_bp.route('/generated_pdfs/<path:filename>')
def serve_generated_pdfs(filename):
    pdfs_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'registration_backend', 'generated_pdfs')
    if not os.path.exists(pdfs_dir):
        pdfs_dir = 'generated_pdfs'
    return send_from_directory(pdfs_dir, filename)

# --- SPA FALLBACK (Handled by main.py for /admin and /admin-login) ---

# --- STARTUP ---

def wait_for_db():
    for attempt in range(1, 4):
        try:
            conn = get_db()
            conn.close()
            logger.info("Database connection established")
            return True
        except Exception as e:
            logger.warning(f"DB check attempt {attempt}/3: {e}")
            time.sleep(1)
    logger.warning("Database connection deferred.")
    return False

# When run directly (dev server), wait for DB then start Flask
if __name__ == '__main__':
    wait_for_db()
    try:
        from gunicorn.app.wsgiapp import WSGIApplication
        WSGIApplication().run()
    except ImportError:
        logger.warning("gunicorn not installed, using Flask dev server (single-threaded, dev only)")
else:
    wait_for_db()
